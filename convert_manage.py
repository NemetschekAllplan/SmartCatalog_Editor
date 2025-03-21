#!/usr/bin/python3
# -*- coding: utf-8 -*
import json
import os.path
import shutil
import sqlite3
import time
import zipfile
from typing import Union

import dbf
import openpyxl
from lxml import etree
from openpyxl import Workbook

from allplan_manage import AllplanDatas, Creation, AttributeDatas
from hierarchy import *
from tools import afficher_message as msg, find_new_title, read_file_to_list, qm_check, excel_load_workbook
from tools import convertir_bytes, find_folder_path, get_value_is_valid, xml_load_root, application_title
from tools import open_file, find_filename, convertir_nom_fichier_correct, settings_save_value
from translation_manage import *


def a___________________detection______():
    pass


class BddTypeDetection(QObject):

    def __init__(self):
        super().__init__()

        self.error_message = ""
        self.bdd_type = ""
        self.bdd_title = ""
        self.file_path = ""

    def search_bdd_type(self, file_path: str):

        error_txt = self.tr("Erreur")

        bible_txt = self.tr("Bible externe")

        if file_path == "":
            error_message = self.tr("Aucun chemin défini.")

            self.error_message = f"{error_txt} -- {error_message}"

            return False

        if file_path.lower().endswith(".txt"):
            part1 = self.tr("Ce format de fichier n'est pas pris en charge.")
            part2 = self.tr("Si ce fichier est un fichier csv, il est nécessaire de changer l'extention.")

            error_message = f"{part1}\n {part2}"

            self.error_message = f"{error_txt} -- {error_message}"

            return False

        # ---------------------------------------
        # HTTPS
        # ---------------------------------------

        if file_path.startswith("https"):

            self.file_path = file_path

            # ---------------------------------------
            # SYNERMI
            # ---------------------------------------

            if bdd_type_synermi.lower() in file_path.lower():
                self.bdd_type = bdd_type_synermi
                self.bdd_title = "Synermi"
                return True

            # ---------------------------------------
            # CAPMI
            # ---------------------------------------

            if bdd_type_capmi.lower() in file_path.lower():
                self.bdd_type = bdd_type_capmi
                self.bdd_title = "Capmi"
                return True

            # ---------------------------------------
            # PROGEMI
            # ---------------------------------------

            if bdd_type_progemi.lower() in file_path.lower():
                self.bdd_type = bdd_type_progemi
                self.bdd_title = "Progemi"
                return True

            # ---------------------------------------
            # Excel
            # ---------------------------------------

            if ".xlsx" in file_path.lower() and "/" in file_path:
                if self.excel_file():
                    return True

            self.bdd_type = bdd_type_extern
            self.bdd_title = bible_txt

            return True

        # ---------------------------------------

        title = find_filename(file_path).title()

        if title == "":
            error_message = self.tr("Une erreur est survenue.")

            self.error_message = f"{error_txt} -- {error_message}"

            return False

        temp = file_path.upper()

        self.file_path = file_path

        error_message = self.tr("Cette base de données n'a pas été reconnue.")

        # ---------------------------------------
        # KUKAT
        # ---------------------------------------

        if temp.endswith(".KAT"):
            self.bdd_type = bdd_type_kukat
            self.bdd_title = f"{bible_txt} - Kukat"
            return True

        # ---------------------------------------
        # EXCEL
        # ---------------------------------------

        if temp.endswith(".XLSX"):
            if self.excel_file():
                return True

        # ---------------------------------------
        # CSV
        # ---------------------------------------

        if temp.endswith(".CSV"):
            self.bdd_type = bdd_type_excel
            self.bdd_title = title
            return True

        # ---------------------------------------
        # MXDB
        # ---------------------------------------

        if temp.endswith(".MXDB"):
            self.bdd_type = bdd_type_mxdb
            self.bdd_title = title
            return True

        # ---------------------------------------
        # EXCEL / CSV
        # ---------------------------------------

        if temp.endswith(".XPWE"):
            self.bdd_type = bdd_type_xpwe
            self.bdd_title = title
            return True

        # ---------------------------------------
        # XML
        # ---------------------------------------

        if temp.endswith(".XML"):

            try:

                root = xml_load_root(file_path=file_path)

                if not isinstance(root, etree._Element):
                    return False

                brand_txt = root.tag

                # ---------------------------------------
                # Nevaris
                # ---------------------------------------

                if brand_txt == "{six.xsd}Documento":
                    self.bdd_type = bdd_type_team_system_xml
                    self.bdd_title = title
                    return True

                # ---------------------------------------
                # Nevaris
                # ---------------------------------------

                if brand_txt == '{urn:schemas-microsoft-com:office:spreadsheet}Workbook':

                    try:

                        first_row = root.find(path='.//ss:Row[2]',
                                              namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'})

                        if first_row is None:
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        style_id = first_row.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}StyleID', '')

                        if not isinstance(style_id, str):
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        if 'NEVARIS_STYLE_H1' != style_id:
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        datas = first_row.findall(path=".//ss:Data",
                                                  namespaces={'ss': 'urn:schemas-microsoft-com:office:spreadsheet'})

                        if len(datas) < 2:
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        bdd_type = datas[0].text

                        if bdd_type != "Raumelement":
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                        title = datas[1].text

                        if not isinstance(title, str):
                            self.error_message = f"{error_txt} -- {error_message}"
                            return False

                    except Exception:
                        self.error_message = f"{error_txt} -- {error_message}"
                        return False

                    self.bdd_type = bdd_type_nevaris
                    self.bdd_title = title
                    return True

                # ---------------------------------------
                # SMARTCATALOG
                # ---------------------------------------

                if brand_txt == "AllplanCatalog":

                    check_version = root.find("Node")

                    if check_version is None:
                        self.error_message = f"{error_txt} -- {error_message}"
                        return False

                    self.bdd_type = bdd_type_xml
                    self.bdd_title = title
                    return True

                # ---------------------------------------
                # SMARTCATALOG
                # ---------------------------------------

                check_version = root.find("Folder")

                if check_version is not None:

                    if brand_txt not in bdd_icons_dict:
                        self.bdd_type = bdd_type_extern
                    else:
                        self.bdd_type = brand_txt

                    self.bdd_title = brand_txt.title()
                    return True

                # ---------------------------------------
                # SMARTCATALOG - Extern
                # ---------------------------------------

                check_version = root.find("Dossier")

                if check_version is not None:

                    # ---------------------------------------
                    # GIMI
                    # --------------------------------------

                    if brand_txt.upper() == "GIMI":
                        self.bdd_type = bdd_type_gimi
                        self.bdd_title = "Gimi"
                        return True

                    # ---------------------------------------
                    # Other
                    # --------------------------------------

                    if brand_txt not in bdd_icons_dict:
                        self.bdd_type = bdd_type_extern
                    else:
                        self.bdd_type = brand_txt

                    self.bdd_title = brand_txt.title()
                    return True

                # ---------------------------------------
                # SMARTCATALOG - Extern - Old
                # ---------------------------------------

                check_version = root.find("Dossier")

                if check_version is not None:

                    if brand_txt not in bdd_icons_dict:
                        self.bdd_type = bdd_type_extern
                    else:
                        self.bdd_type = brand_txt

                    self.bdd_title = brand_txt
                    return True

                self.error_message = f"{error_txt} -- {error_message}"

                return False

            except Exception:

                self.error_message = f"{error_txt} -- {error_message}"

            return False

        # ---------------------------------------
        # ARA
        # ---------------------------------------

        if temp.endswith(".ARA"):

            file_path = unpack_ara_file(file_path=file_path)

            if file_path == "":
                return False

            self.file_path = file_path

            temp = file_path.upper()

        # ---------------------------------------
        # DBF
        # ---------------------------------------
        if temp.endswith(".DBF"):

            try:

                with dbf.Table(filename=file_path) as table:

                    if len(table) == 0:
                        self.error_message = f"{error_txt} -- {error_message}"

                        return False

                    liste_colonnes = table.field_names

                    if "VWKTX" not in liste_colonnes:
                        self.error_message = f"{error_txt} -- {error_message}"

                        return False

                    record = table.first_record
                    item_type: str = convertir_bytes(record.VWTPU)
                    code_item_type: str = convertir_bytes(record.VWTYP)
                    title = convertir_bytes(record.VWKTX)

                    title = convertir_nom_fichier_correct(title).title()

                    if code_item_type != "X":
                        self.error_message = f"{error_txt} -- {error_message}"
                        return False

                    # ---------------------------------------
                    # ALLMETRE - EURICIEL
                    # ---------------------------------------

                    if item_type == "Projekt":
                        self.bdd_type = bdd_type_allmetre_e
                        self.bdd_title = title
                        return True

                    # ---------------------------------------
                    # ALLMETRE - AJSOFT
                    # ---------------------------------------

                    if item_type == "Pos.":
                        self.bdd_type = bdd_type_allmetre_a
                        self.bdd_title = title
                        return True

                    for record in table:

                        if dbf.is_deleted(record):
                            continue

                        code_item_type: str = convertir_bytes(record.VWTYP)

                        # ---------------------------------------
                        # BCM - MATERIAL
                        # ---------------------------------------

                        if code_item_type == "E" or code_item_type == "T":
                            self.bdd_type = bdd_type_bcm
                            self.bdd_title = title
                            return True

                        # ---------------------------------------
                        # BCM - COMPONENT
                        # ---------------------------------------

                        if code_item_type == "L":
                            self.bdd_type = bdd_type_bcm_c
                            self.bdd_title = title
                            return True

            except Exception as error:
                print(f"convert_manage -- search_bdd_type -- error : {error}")
                pass

            self.error_message = f"{error_txt} -- {error_message}"

            return False

        if temp.endswith(".FIC") or temp.endswith(".NDX") or temp.endswith(".MMO"):

            if title.upper() in ["CM", "ST", "ARTICLES", "PARA"]:

                path_gimi = os.path.dirname(file_path)

                path_gimi = path_gimi.replace("/", "\\")

                if not path_gimi.endswith("\\"):
                    path_gimi += "\\"

                settings_save_value(library_setting_file, "path_gimi", path_gimi)

                # ---------------------------------------
                # GIMI
                # ---------------------------------------

                self.bdd_type = bdd_type_gimi
                self.bdd_title = f"{bible_txt} - {bdd_type_gimi}"
                return True

            return False

        dict_favoris_allplan = get_favorites_allplan_dict()

        for extension, nom_favoris in dict_favoris_allplan.items():

            if file_path.endswith(extension):
                self.bdd_type = bdd_type_fav
                self.bdd_title = title
                return True

        self.error_message = f"{error_txt} -- {error_message}"
        return False

    def excel_file(self) -> bool:

        try:

            workbook = excel_load_workbook(file_path=self.file_path)

            if not isinstance(workbook, openpyxl.Workbook):
                return False

            sheet = workbook.active

            test_1 = sheet.cell(1, 7).value

            if isinstance(test_1, str):

                if 'INCIDENZA CATEGORIE\n(%)' in test_1:
                    self.bdd_type = bdd_type_team_system_xlsx
                    self.bdd_title = workbook.sheetnames[0]
                    return True

            # -------------

            test_2 = sheet.cell(1, 1).value

            if isinstance(test_2, str):
                if test_2 == "Teilleistungsnummer":
                    self.bdd_type = bdd_type_nevaris_xlsx
                    self.bdd_title = workbook.sheetnames[0]
                    return True

            # -------------

            self.bdd_type = bdd_type_excel

            if self.file_path.startswith("https"):

                path_split = self.file_path.split("/")

                for part in path_split:

                    part = part.lower()

                    if ".xlsx" not in part:
                        continue

                    title = part.replace(".xlsx", "").strip()

                    self.bdd_title = title.title()
                    return True

            else:
                self.bdd_title = find_filename(file_path=self.file_path).title()
                return True

            self.bdd_title = workbook.sheetnames[0]
            return True

        except Exception as error:
            print(f"convert_manage -- BddTypeDetection -- excel_file -- error: {error}")
            pass

        return False


def unpack_ara_file(file_path: str) -> str:
    # -------------------------
    # Verification file exists
    # -------------------------

    if not isinstance(file_path, str):
        print(f"conver_manage -- unpack_ara_file -- not isinstance(file_path, str)")
        return ""

    if not os.path.exists(file_path):
        print(f"conver_manage -- unpack_ara_file -- not os.path.exists(file_path)")
        return ""

    # -------------------------
    # search folder path
    # -------------------------

    folder_path = find_folder_path(file_path)

    if folder_path == "":
        print(f"conver_manage -- unpack_ara_file -- folder_path == empty")
        return ""

    # -------------------------
    # search filename
    # -------------------------

    file_name = find_filename(file_path=file_path)

    file_name = file_name.strip()

    if file_name == "":
        print(f"conver_manage -- unpack_ara_file -- file_name == empty")
        return ""

    # -------------------------
    # define others paths
    # -------------------------

    export_path = f"{asc_export_path}{file_name}\\"

    json_file_path = f"{export_path}{file_name}.ini"

    export_path_bak = f"{asc_export_path}{file_name}_bak"

    dbf_path = f"{export_path}VW1.DBF"

    # -------------------------
    # search size ARA file
    # -------------------------

    try:
        file_size = os.path.getsize(file_path)

    except OSError as error:
        print(f"conver_manage -- unpack_ara_file -- error delete : {error}")
        return ""

    # -------------------------
    # search old size
    # -------------------------

    file_size_old = 0

    if os.path.exists(json_file_path):

        try:

            with open(json_file_path, 'r', encoding="Utf-8") as file:

                file_size_old = json.load(file)

        except Exception as error:
            print(f"conver_manage -- unpack_ara_file -- error json read : {error}")
            file_size_old = 0
            pass

    # -------------------------
    # current ARA file and old ARA file are same
    # -------------------------

    if file_size == file_size_old and os.path.exists(dbf_path):
        return dbf_path

    # -------------------------
    # Delete backup folder
    # -------------------------

    if os.path.exists(export_path_bak):

        try:
            shutil.rmtree(export_path_bak)
        except Exception as error:
            print(f"conver_manage -- unpack_ara_file -- error delete : {error}")
            return ""

    # -------------------------
    # Rename current folder to backup folder
    # -------------------------

    if os.path.exists(export_path):

        try:
            os.rename(export_path, export_path_bak)
        except Exception as error:
            print(f"conver_manage -- unpack_ara_file -- error rename : {error}")
            return ""

    # -------------------------
    # Create new folder
    # -------------------------

    try:
        os.makedirs(export_path)

    except Exception as error:
        print(f"conver_manage -- unpack_ara_file -- error make dir : {error}")
        return ""

    # -------------------------
    # Unzip
    # -------------------------

    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(export_path)

    except Exception as error:
        print(f"conver_manage -- unpack_ara_file -- error zip : {error}")
        return ""

    # -------------------------
    # search dbf file
    # -------------------------

    if not os.path.exists(dbf_path):
        print(f"conver_manage -- unpack_ara_file -- not os.path.exists(dbf_path)")
        return ""

    # -------------------------
    # Write json file
    # -------------------------

    try:
        with open(json_file_path, 'w', encoding="Utf-8") as file:

            json.dump(file_size, file, ensure_ascii=False, indent=2)

    except Exception as error:
        print(f"conver_manage -- unpack_ara_file -- error json write : {error}")
        return "False"

    # -------------------------
    # Unpack is ok
    # -------------------------

    return dbf_path


class ConvertTemplate(QObject):
    loading_completed = pyqtSignal(QStandardItemModel, list, list, list)
    number_error_signal = pyqtSignal(list)
    errors_signal = pyqtSignal(list)

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):

        super().__init__()

        # --------------
        # Allplan
        # --------------

        self.allplan: AllplanDatas = allplan
        self.creation: Creation = self.allplan.creation

        if conversion:
            self.allplan.creation.attributes_datas.clear()

        # --------------
        # Model
        # --------------

        self.cat_model = QStandardItemModel()

        self.cat_model.setHorizontalHeaderLabels([bdd_title, self.tr("Description")])

        self.qs_root = self.cat_model.invisibleRootItem()
        self.qs_root.setData(folder_code, user_data_type)

        # --------------

        self.material_list = list()
        self.material_upper_list = list()
        self.link_list = list()
        self.material_with_link_list = list()

        # --------------
        # Variables BDD
        # --------------

        self.file_path = file_path
        self.bdd_title = bdd_title
        self.conversion = conversion

        # ----

        self.expanded_list = list()
        self.selected_list = list()

        self.errors_list = list()

        self.number_error_list = list()

        # ----

        self.counter = time.perf_counter()

    def start_loading(self):

        self.cat_model.beginResetModel()

        print(f"convert_manage -- {type(self)}  -- read {self.file_path} ---------- ")

    def end_loading(self):

        self.cat_model.endResetModel()
        self.loading_completed.emit(self.cat_model, self.expanded_list, self.selected_list, self.number_error_list)

        print(f"DB loaded in : {time.perf_counter() - self.counter} ms")

        # -----------------

        if len(self.errors_list) != 0:
            self.errors_signal.emit(self.errors_list)

        # -----------------

        if len(self.number_error_list) != 0:
            self.number_error_signal.emit(self.number_error_list)


def a___________________bcm_component______():
    pass


class BcmArticleComponent:

    def __init__(self, type_ele: str, cid_index: str, pid_index: str, srt_index: int, code: str, desc: str):
        super().__init__()

        # ---------
        # Required
        # ---------

        self.type_ele = type_ele

        self.cid_index = cid_index
        self.pid_index = pid_index
        self.srt_index = srt_index

        self.code = code
        self.desc = desc

        # ---------
        # Optional
        # ---------

        self.full_text = ""

        self.unit_value = ""

        self.trade_value = ""

        self.price = ""

        self.formula = ""
        self.materiaux_dyn = ""
        self.formula_obj = ""
        self.quantity = ""


class ConvertBcmComposants(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------
        self.root_code = "Root"

        self.keys_convert = {"X": self.root_code,
                             "E": folder_code,
                             "T": folder_code,
                             "O": folder_code,
                             "L": folder_code,
                             "P": component_code}

        # --------------
        # Variables datas
        # --------------

        self.prices_dict = dict()
        self.pid_dict = dict()
        self.cid_root = ""

        # --------------

    def run(self):

        self.start_loading()

        # ----------------------

        self.get_all_prices()

        # ----------------------
        # Read DB
        # ----------------------

        try:
            with dbf.Table(filename=self.file_path) as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    self.bcm_load_record_datas(record=record)

        except Exception as error:
            print(f"convert_manage -- ConvertBcmComposants -- run -- erreur : {error}")
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        # ----------------------
        # Load DB
        # ----------------------

        if self.cid_root == "":
            print(f"convert_manage -- ConvertBcmComposants -- run -- self.cid_root is empty")
            self.errors_list.append(f"run -- cid_root is empty")
            self.end_loading()
            return False

        if self.cid_root not in self.pid_dict:
            print(f"convert_manage -- ConvertBcmComposants -- run -- self.cid_root not in self.pid_dict")
            self.errors_list.append(f"run -- cid_root not in self.pid_dict")
            self.end_loading()
            return False

        children_articles = self.pid_dict.get(self.cid_root)

        if not isinstance(children_articles, list):
            print(f"convert_manage -- ConvertBcmComposants -- run -- ot isinstance(children_articles, list)")
            self.errors_list.append(f"run -- not isinstance(children_articles, list)")
            self.end_loading()
            return False

        self.bcm_load_hierarchy(children_articles=children_articles, qs_parent=self.qs_root)

        self.end_loading()

        return True

    def bcm_load_record_datas(self, record) -> None:

        # ----------------------
        # Type - Required
        # ----------------------

        key_type = convertir_bytes(record.VWTYP)

        if not isinstance(key_type, str):
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- not isinstance(key_type, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(key_type, str)")
            return

        type_ele = self.keys_convert.get(key_type, "")

        if type_ele == "":
            return

        # ----------------------
        # CID - Required
        # ----------------------

        cid_index = convertir_bytes(record.VWCID)

        if not isinstance(cid_index, str):
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- not isinstance(cid_index, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(cid_index, str)")
            return

        if cid_index == "":
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- cid_index is empty")
            self.errors_list.append(f"bcm_load_record_datas -- cid_index is empty")
            return

        if type_ele == self.root_code:
            self.cid_root = cid_index
            return

        # ----------------------
        # PID - Required
        # ----------------------

        pid_index = convertir_bytes(record.VWPID)

        if not isinstance(pid_index, str):
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- not isinstance(pid_index, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(pid_index, str)")
            return

        if pid_index == "" and type_ele != self.root_code:
            pid_index = self.cid_root

        # ----------------------
        # SRT - Required
        # ----------------------

        srt_index = convertir_bytes(record.VWSRT)

        if not isinstance(srt_index, int):
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- not isinstance(srt_index, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(srt_index, str)")
            return

        if srt_index == -1:
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- srt_index == -1")
            self.errors_list.append(f"bcm_load_record_datas -- srt_index == -1")
            return

        try:

            srt_index = f"{srt_index:010d}"

        except Exception as error:

            print(f"convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- error : {error}")
            self.errors_list.append(f"bcm_load_record_datas -- format srt index : {error}")
            return

        # ----------------------
        # Code element - Required
        # ----------------------

        code = convertir_bytes(record.VWPNR)

        if not isinstance(code, str):
            code = ""

        # ----------------------
        # Description - Required
        # ----------------------

        desc = convertir_bytes(record.VWKTX)

        if not isinstance(desc, str):
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- not isinstance(desc, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(desc, str)")
            return

        if code == "" and desc != "":
            code = desc

        # ----------------------
        # Article creation
        # ----------------------

        article = BcmArticleComponent(type_ele=type_ele,
                                      cid_index=cid_index,
                                      pid_index=pid_index,
                                      srt_index=srt_index,
                                      code=code,
                                      desc=desc)

        # ----------------------

        if type_ele == folder_code:
            self.bcm_add_article_to_dict(pid_index=pid_index, article=article)
            return

        # ----------------------
        # Full_Text - Optional
        # ----------------------

        full_text = convertir_bytes(record.VWLTX)

        if not isinstance(full_text, str):
            print("convert_manage -- ConvertBcmComposants -- bcm_load_record_datas -- not isinstance(full_text, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(full_text, str)")
            return None

        article.full_text = full_text

        # ----------------------
        # Trade - Optional
        # ----------------------

        trade_value = convertir_bytes(record.VWGEW)

        if isinstance(trade_value, str):
            article.trade_value = trade_value

        # ----------------------
        # Unit - Optional
        # ----------------------

        unit_value = convertir_bytes(record.VWDIM)

        if not isinstance(unit_value, str):
            unit_value = ""

        article.unit_value = unit_value

        # ----------------------
        # Price - Optional
        # ----------------------

        article.price = self.prices_dict.get(cid_index, "")

        # ----------------------
        # Formula - Optional
        # ----------------------

        formule_tps = convertir_bytes(record.VWATR)

        if not isinstance(formule_tps, str):
            formula = materiaux_dyn = formula_obj = quantity = ""
        else:
            formula, materiaux_dyn, formula_obj, quantity = self.allplan.convertir_formule_bdd(formule_tps)

        article.formula = formula
        article.materiaux_dyn = materiaux_dyn
        article.formula_obj = formula_obj
        article.quantity = quantity

        # ----------------------

        self.bcm_add_article_to_dict(pid_index=pid_index, article=article)

        # ----------------------

    def bcm_add_article_to_dict(self, pid_index: str, article):

        cid_list = self.pid_dict.get(pid_index)

        if not isinstance(cid_list, list):
            self.pid_dict[pid_index] = [article]
        else:
            cid_list.append(article)

    def bcm_load_hierarchy(self, children_articles: list, qs_parent: QStandardItem):

        children_articles.sort(key=lambda x: x.srt_index)

        for article in children_articles:

            if not isinstance(article, BcmArticleComponent):
                print(f"convert_manage -- ConvertBcmComposants -- bcm_load_hierarchy -- "
                      f"not isinstance(article, BcmComponent)")

                self.errors_list.append(f"bcm_load_hierarchy -- not isinstance(article, BcmArticleComponent)")
                return False

            # --------------
            # Component creation
            # --------------

            if article.type_ele == component_code:

                qs_component_list = self.allplan.creation.component_line(value=article.code,
                                                                         description=article.desc,
                                                                         tooltips=False)

                if len(qs_component_list) != col_cat_count:
                    print(f"convert_manage -- ConvertBCmComponent -- bcm_load_hierarchy -- "
                          f"len(qs_component_list) != col_cat_count")
                    self.errors_list.append(f"bcm_load_hierarchy -- len(qs_component_list) != col_cat_count")
                    continue

                qs_component: Component = qs_component_list[col_cat_value]

                self.attributes_add(qs_value=qs_component, article=article)

                qs_parent.appendRow(qs_component_list)

                continue

            # --------------
            # Folder creation
            # --------------

            qs_folder_list = self.creation.folder_line(value=article.code,
                                                       description=article.desc,
                                                       tooltips=False)

            if len(qs_folder_list) != col_cat_count:
                print(f"convert_manage -- ConvertBcmComposants -- bcm_load_hierarchy -- "
                      f"len(qs_folder_list) != col_cat_count")
                self.errors_list.append(f"bcm_load_hierarchy -- len(qs_folder_list) != col_cat_count")
                continue

            qs_folder: Folder = qs_folder_list[col_cat_value]

            children_articles = self.pid_dict.get(article.cid_index)

            if not isinstance(children_articles, list):
                print(f"convert_manage -- ConvertBcmComposants -- bcm_load_hierarchy -- "
                      f"not isinstance(children_articles, list)")
                self.errors_list.append(f"bcm_load_hierarchy -- not isinstance(children_articles, list)")
                continue

            # --------------
            # Children creation
            # --------------

            if len(children_articles) != 0:
                self.bcm_load_hierarchy(children_articles=children_articles, qs_parent=qs_folder)

            qs_parent.appendRow(qs_folder_list)

    @staticmethod
    def a___________________bcm_attribute__________________():
        pass

    def attributes_add(self, qs_value: QStandardItem, article: BcmArticleComponent):

        # ---------------
        # Unit @202@
        # ---------------

        if article.unit_value != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.unit_value, number_str="202"))

        # ---------------
        # Price @203@
        # ---------------

        if article.price != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.price, number_str="203"))

        # ---------------
        # Full text @208@
        # ---------------

        if article.full_text != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.full_text, number_str="208"))

        # ---------------
        # Trade @209@
        # ---------------

        if article.trade_value != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.trade_value, number_str="209"))

        # ---------------

        if isinstance(qs_value, Material):
            return

        # ---------------
        # Object_filter @76@
        # ---------------

        if article.formula_obj != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.formula_obj, number_str="76"))

        # ---------------
        # Material_dyn @96@
        # ---------------

        if article.materiaux_dyn != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.materiaux_dyn, number_str="96"))

        # ---------------
        # Piece @215@
        # ---------------

        if article.quantity != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.quantity, number_str="215"))

        # ---------------
        # Formula @267@
        # ---------------

        if article.formula == "":
            article.formula = self.allplan.recherche_formule_defaut(unit=article.unit_value)

        qs_value.appendRow(self.creation.attribute_line(value=article.formula, number_str="267"))

    @staticmethod
    def a___________________bdd_bcm_price__________________():
        pass

    def get_all_prices(self):

        file_name = find_filename(file_path=self.file_path)

        file_name_new = file_name.lower().replace("vw", "pr")

        folder_path = find_folder_path(file_path=self.file_path)

        file_path = f"{folder_path}{file_name_new}.dbf"

        if not os.path.exists(file_path):
            return

        try:
            with dbf.Table(filename=file_path) as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    price_float = convertir_bytes(record.PREPB)

                    if not isinstance(price_float, float):
                        continue

                    if price_float == 0:
                        continue

                    price_str = f"{price_float:.2f}"

                    current_code = convertir_bytes(record.PRIDPOS)

                    if not isinstance(current_code, str):
                        continue

                    if current_code in self.prices_dict:
                        continue

                    self.prices_dict[current_code] = price_str

        except Exception as error:
            print(f"convert_manage -- ConvertBcmComposants -- get_all_prices -- Erreur : {error}")
            self.errors_list.append(f"get_all_prices -- {error}")
            return

    @staticmethod
    def a___________________end__________________():
        pass


def a___________________bcm_material______():
    pass


class BcmArticleMaterial:

    def __init__(self, type_ele: str, cid_index: str, pid_index: str, srt_index: int, code: str, desc: str):

        super().__init__()

        # ---------
        # Required
        # ---------

        self.type_ele = type_ele

        self.cid_index = cid_index
        self.pid_index = pid_index
        self.srt_index = srt_index

        self.code = code
        self.desc = desc

        # ---------
        # Optional
        # ---------

        self.full_text = ""

        self.unit_value = ""

        self.trade_value = ""

        self.price = ""

        self.formula = ""
        self.materiaux_dyn = ""
        self.formula_obj = ""
        self.quantity = ""

        # ---------

        self.children_article_list = list()

        self.children_type_list = list()

        # ---------

        self.grouped = False

    def analyse_children(self):

        # ---------------
        # Sort items
        # ---------------

        if len(self.children_article_list) != 0:
            self.children_article_list.sort(key=lambda article: article.srt_index)

        # ----------------

        # ---------------
        # Check analyze needed
        # ---------------

        if self.type_ele == component_code or self.type_ele == link_code or self.type_ele == "Comment":
            return

        # ---------------
        # Analyze Children
        # ---------------

        for article_child in self.children_article_list:
            article_child: BcmArticleMaterial
            article_child.analyse_children()

        self.get_children_type()

        # ---------------- Debug test

        # if self.code == 'ATTRIB_PROJET':
        #     print("b")

        # ----------------

        if len(self.children_type_list) == 0:
            if self.type_ele == folder_code:
                self.type_ele = material_code
            return

            # ----------------

        folder_exist = folder_code in self.children_type_list

        material_exist = material_code in self.children_type_list

        component_exist = (component_code in self.children_type_list or link_code in self.children_type_list)

        comment_exist = "Comment" in self.children_type_list

        # ----------------

        if not folder_exist and not material_exist and not component_exist:
            # ============================================ (1)
            # Folder                        (Current - empty)
            # ============================================

            if comment_exist:
                self.change_all_comment_to(type_ele=folder_code)
            return

        # ----------------

        if folder_exist and not material_exist and not component_exist:
            # ============================================ (2)
            # Folder                        (Current)
            # --- Folder                    (child)
            # --- Folder                    (child)
            # ============================================

            if comment_exist:
                self.change_all_comment_to(type_ele=folder_code)

            return

        # ----------------

        if not folder_exist and material_exist and not component_exist:
            # ============================================ (3)
            # Folder                        (Current)
            # --- Material                  (child)
            # --- Material                  (child)
            # ============================================

            if self.type_ele == "Root":
                self.move_every_material_to_new_folder()
                type_ele = folder_code
            else:
                type_ele = material_code

            if comment_exist:
                self.change_all_comment_to(type_ele=type_ele)

            return

        # ----------------

        if not folder_exist and not material_exist and component_exist:
            # ============================================ (4)
            # Folder                        (Current) -> This is a material
            # --- Component                 (child)
            # --- Component                 (child)
            # ============================================
            self.type_ele = material_code

            if comment_exist:
                self.change_all_comment_to(type_ele=component_code)

            return

        # ----------------

        if folder_exist and material_exist and not component_exist:
            # ============================================ (5)
            # Folder                        (Current)
            # --- Folder                    (child)
            # --- Material                  (child) -> move every material into a new folder
            # ============================================

            self.move_every_material_to_new_folder()

            if comment_exist:
                self.change_all_comment_to(type_ele=folder_code)

            return

        # ----------------

        if folder_exist and not material_exist and component_exist:
            # ============================================ (6)
            # Folder                        (Current)
            # --- Folder                    (child) -> this a material
            # --- Component                 (child) -> need create Material and move components into
            # ============================================

            if self.type_ele == "Root":
                self.move_all_component_to_folder()

                if comment_exist:
                    self.change_all_comment_to(type_ele=folder_code)

                return

            # self.change_all_folders_to_material()

            self.move_all_component_to_folder(code=self.code, desc=self.desc)

            if comment_exist:
                self.change_all_comment_to(type_ele=folder_code)

            return

        # ------------------------------------------------------------------------------------------------

        if not folder_exist and material_exist and component_exist:
            # ============================================ (7)
            # Folder                        (Current)
            # --- Material                  (child)
            # --- Component                 (child) -> need create Material and move components into
            # ============================================

            self.move_all_components_to_new_material()

            return

        # ------------------------------------------------------------------------------------------------

        if folder_exist and material_exist and component_exist:
            # ============================================ (8)
            # Folder                        (Current)
            # --- Folder                    (child) -> ok
            # --- Material                  (child) -> move every material into a new folder
            # --- Component                 (child) -> need create Material and move components into
            # ============================================

            self.move_all_components_to_new_material()
            self.move_every_material_to_new_folder()

            if comment_exist:
                self.change_all_comment_to(type_ele=folder_code)

            return

        print(f"{folder_exist} + {material_exist} + {component_exist}")

    @staticmethod
    def a___________________type_list__________________():
        pass

    def get_children_type(self):
        self.children_type_list = [article_child.type_ele for article_child in self.children_article_list]

    @staticmethod
    def a___________________change_type__________________():
        pass

    def change_all_folders_to_material(self):
        """
        change all folder to material
        :return:
        """

        for article_child in self.children_article_list:
            article_child: BcmArticleMaterial

            if article_child.type_ele == folder_code:
                article_child.type_ele = material_code

    def change_all_comment_to(self, type_ele: str):

        for article_child in self.children_article_list:
            article_child: BcmArticleMaterial

            if article_child.type_ele == "Comment":
                self.change_comment_to(article_child=article_child, type_ele=type_ele)

    def change_comment_to(self, article_child, type_ele: str):
        article_child.type_ele = type_ele

        if self.code == "":
            code = "--------"
        else:
            code = self.code

        article_child.code = f" -------- {code} --------"

        if article_child.desc == "":
            article_child.desc = "-" * 50

        if type_ele == component_code:
            article_child.formula = "0"

    @staticmethod
    def a___________________move__________________():
        pass

    def move_every_material_to_new_folder(self):
        """
        Move all materials into new folder
        :return:
        """

        for index_row, article_child in enumerate(self.children_article_list):

            article_child: BcmArticleMaterial

            if article_child.type_ele == "Comment":
                self.change_comment_to(article_child=article_child, type_ele=folder_code)

            if article_child.type_ele != material_code:
                continue

            article_folder = BcmArticleMaterial(type_ele=folder_code,
                                                cid_index=article_child.cid_index,
                                                pid_index=self.cid_index,
                                                srt_index=article_child.srt_index,
                                                code=article_child.code,
                                                desc=article_child.desc)

            article_folder.children_article_list.append(article_child)
            article_folder.get_children_type()

            self.children_article_list[index_row] = article_folder

        self.get_children_type()

    def move_all_components_to_new_material(self, copy_sub_component=True):
        """
        Move all components/links/Comment (child) into new material (child)
        :return:
        """

        cid_index = f"{self.cid_index}_M"

        article_material = BcmArticleMaterial(type_ele=material_code,
                                              cid_index=cid_index,
                                              pid_index=self.cid_index,
                                              srt_index=-1,
                                              code=self.code,
                                              desc=self.desc)

        children_article_delete = list()

        for index_row, article_child in enumerate(self.children_article_list):

            article_child: BcmArticleMaterial

            if article_child.type_ele == folder_code or article_child.type_ele == material_code:

                # copy all component of all folders and materials
                if copy_sub_component:
                    self.copy_sub_components_to_materials(article_original=article_child,
                                                          article_target=article_material)
                continue

            article_material.srt_index = article_child.srt_index

            children_article_delete.append(index_row)

            article_child.pid_index = cid_index

            if article_child.type_ele == "Comment":
                self.change_comment_to(article_child=article_child, type_ele=component_code)

            article_material.children_article_list.append(article_child)

        for index_row in reversed(children_article_delete):
            self.children_article_list.pop(index_row)

        article_material.get_children_type()

        self.children_article_list.insert(0, article_material)

        self.grouped = copy_sub_component

        self.get_children_type()

    def copy_sub_components_to_materials(self, article_original, article_target):

        if not isinstance(article_original, BcmArticleMaterial) or not isinstance(article_target, BcmArticleMaterial):
            return

        for article_child in article_original.children_article_list:

            if not isinstance(article_child, BcmArticleMaterial):
                continue

            type_ele = article_child.type_ele

            if type_ele == folder_code or type_ele == material_code:
                self.copy_sub_components_to_materials(article_original=article_child, article_target=article_target)
                continue

            if article_child not in article_target.children_article_list:
                article_target.children_article_list.append(article_child)

    def move_all_component_to_folder(self, code="Root", desc="Root"):

        component_article_list = list()

        for article_child in reversed(self.children_article_list):

            if not isinstance(article_child, BcmArticleMaterial):
                continue

            type_ele = article_child.type_ele

            if type_ele == component_code or type_ele == link_code:
                component_article_list.append(article_child)

                self.children_article_list.remove(article_child)

        if len(component_article_list) == 0:
            return

        cid_index = component_article_list[0].cid_index

        folder_article = BcmArticleMaterial(type_ele=folder_code,
                                            cid_index=f"{cid_index}_F",
                                            pid_index=self.cid_index,
                                            srt_index="",
                                            code=code,
                                            desc=desc)

        material_article = BcmArticleMaterial(type_ele=material_code,
                                              cid_index=f"{cid_index}_M",
                                              pid_index=f"{cid_index}_F",
                                              srt_index="",
                                              code=code,
                                              desc=desc)

        material_article.children_article_list = component_article_list
        material_article.get_children_type()

        folder_article.children_article_list.append(material_article)
        folder_article.get_children_type()

        self.children_article_list.append(folder_article)
        self.get_children_type()

    @staticmethod
    def a___________________material_group__________________():
        pass

    def create_material_group(self):

        for article_child in self.children_article_list:
            article_child: BcmArticleMaterial

            # ---------------
            # Check analyze needed
            # ---------------

            if article_child.type_ele != folder_code:
                return

            if len(article_child.children_article_list) < 2:
                continue

            if article_child.grouped:
                print(f"article : {article_child.code} a été ignoré lors du regroupement")
                continue

            # ---------------
            # Analyze Children
            # ---------------

            children_article_list = list()

            self.create_material_group_action(article_source=article_child, children_article_list=children_article_list)

            if len(children_article_list) == 0:
                continue

            article_child.create_material_group()

            # ---------------------

            # if article_child.code == 'PLANCHER BETON':
            #     print("b")

            material_article = BcmArticleMaterial(type_ele=material_code,
                                                  cid_index=f"{article_child.cid_index}_M",
                                                  pid_index=article_child.pid_index,
                                                  srt_index=article_child.srt_index,
                                                  code=article_child.code,
                                                  desc=article_child.desc)

            material_article.children_article_list = children_article_list
            material_article.get_children_type()

            if folder_code in article_child.children_type_list:

                folder_article = BcmArticleMaterial(type_ele=folder_code,
                                                    cid_index=f"{article_child.cid_index}_M",
                                                    pid_index=article_child.pid_index,
                                                    srt_index=article_child.srt_index,
                                                    code=article_child.code,
                                                    desc=article_child.desc)

                folder_article.children_article_list.insert(0, material_article)

                article_child.children_article_list.insert(0, folder_article)

                article_child.get_children_type()

            else:

                article_child.children_article_list.insert(0, material_article)

                article_child.get_children_type()

    def create_material_group_action(self, article_source, children_article_list: list):

        for article_child in article_source.children_article_list:
            article_child: BcmArticleMaterial

            child_type = article_child.type_ele

            if child_type == component_code or child_type == link_code:
                children_article_list.append(article_child)
                continue

            if child_type == material_code or child_type == folder_code:
                self.create_material_group_action(article_source=article_child,
                                                  children_article_list=children_article_list)

                continue

            print(f"convert_manage -- BcmArticle -- create_material_group_action -- unknown type {child_type}")

    @staticmethod
    def a___________________end__________________():
        pass


class ConvertBcmMaterial(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.root_key = "X"
        self.element_key = "E"
        self.component_key = "P"
        self.link_key = "V"
        self.comment_key = "B"

        self.root_code = "Root"
        self.comment_code = "Comment"

        self.keys_convert = {self.root_key: self.root_code,
                             self.element_key: folder_code,
                             "T": folder_code,
                             "O": folder_code,
                             "L": folder_code,
                             self.component_key: component_code,
                             self.comment_key: self.comment_code,
                             self.link_key: link_code}

        self.material_child = [component_code, link_code, self.comment_code]

        # --------------
        # Variables datas
        # --------------

        self.prices_dict = dict()

        # --------------
        self.folder_article_dict = dict()
        self.material_article_dict = dict()
        self.material_article_with_link_list = list()
        self.link_dict = dict()

        # --------------

        self.root_article = BcmArticleMaterial(type_ele=self.root_code, cid_index="", pid_index="", srt_index=0,
                                               code="", desc="")

        self.cid_dict = dict()

        # --------------

    def run(self):

        self.start_loading()

        # ----------------------

        self.get_all_prices()

        # ----------------------
        # Read DB
        # ----------------------

        try:
            with dbf.Table(filename=self.file_path, unicode_errors="ignore") as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    self.bcm_load_record_datas(record=record)

        except Exception as error:
            print(f"convert_manage -- ConvertBCmMaterial -- run -- erreur : {error}")
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        # ----------------------
        # Analyse DB
        # ----------------------

        self.root_article.analyse_children()

        # ----------------------
        # Load DB
        # ----------------------

        self.bcm_load_hierarchy(article_parent=self.root_article, qs_parent=self.qs_root)

        # ----------------------
        # Verification
        # ----------------------

        self.verification(qs=self.qs_root)

        # ----------------------
        # Links manage
        # ----------------------

        if len(self.link_list) != 0:
            self.link_verification()

        # ----------------------

        self.end_loading()
        return True

    def bcm_load_record_datas(self, record) -> None:

        # ----------------------
        # Type - Required
        # ----------------------

        key_type = convertir_bytes(record.VWTYP)

        if not isinstance(key_type, str):
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- not isinstance(key_type, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(key_type, str)")
            return

        type_ele = self.keys_convert.get(key_type, "")

        if type_ele == "":
            return

        # ----------------------
        # CID - Required
        # ----------------------

        cid_index = convertir_bytes(record.VWCID)

        if not isinstance(cid_index, str):
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- not isinstance(cid_index, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(cid_index, str)")
            return

        if cid_index == "":
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- cid_index is empty")
            self.errors_list.append(f"bcm_load_record_datas -- cid_index is empty")
            return

        if type_ele == self.root_code:
            self.root_article.cid_index = cid_index
            self.cid_dict[cid_index] = self.root_article
            return

        # if cid_index == "_251G0KUFQQ":
        #     print("1882")

        # ----------------------
        # PID - Required
        # ----------------------

        pid_index = convertir_bytes(record.VWPID)

        if not isinstance(pid_index, str):
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- not isinstance(pid_index, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(pid_index, str)")
            return

        if pid_index == "" and type_ele != self.root_code:
            pid_index = self.root_article.cid_index

        if type_ele == component_code and pid_index == self.root_article.cid_index:
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- component can't be a child of root")
            self.errors_list.append(f"bcm_load_record_datas -- component can't be a child of root")
            return

        # ----------------------
        # CID - Required
        # ----------------------

        article_parent = self.cid_dict.get(pid_index, None)

        if not isinstance(article_parent, BcmArticleMaterial):
            article_parent = BcmArticleMaterial(type_ele=type_ele,
                                                cid_index=cid_index,
                                                pid_index="",
                                                srt_index=-1,
                                                code="",
                                                desc="")

        # ----------------------
        # SRT - Required
        # ----------------------

        srt_index = convertir_bytes(record.VWSRT)

        if not isinstance(srt_index, int):
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- not isinstance(srt_index, str)")
            self.errors_list.append(f"bcm_load_record_datas -- not isinstance(srt_index, str)")
            return

        if srt_index == -1:
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- srt_index == -1")
            self.errors_list.append(f"bcm_load_record_datas --  srt_index == -1")
            return

        # ----------------------
        # Code element - Required
        # ----------------------

        code = convertir_bytes(record.VWCTX)

        if not isinstance(code, str) or code == "":

            code = convertir_bytes(record.VWPNR)

            if not isinstance(code, str):
                code = ""

        # ----------------------
        # Description - Required
        # ----------------------

        desc = convertir_bytes(record.VWKTX)

        if not isinstance(desc, str):
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- not isinstance(desc, str)")
            self.errors_list.append(f"bcm_load_record_datas --  not isinstance(desc, str)")
            return

        if code == "" and desc != "":
            code = desc

        # ----------------------
        # Article creation
        # ----------------------

        if cid_index in self.cid_dict:

            article = self.cid_dict[cid_index]

        else:
            article = BcmArticleMaterial(type_ele=type_ele,
                                         cid_index=cid_index,
                                         pid_index=pid_index,
                                         srt_index=srt_index,
                                         code=code,
                                         desc=desc)

        # ----------------------
        # Full_Text - Optional
        # ----------------------

        full_text = convertir_bytes(record.VWLTX)

        if not isinstance(full_text, str):
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_record_datas -- not isinstance(full_text, str)")
            self.errors_list.append(f"bcm_load_record_datas --  not isinstance(full_text, str)")
            return None

        article.full_text = full_text

        # ----------------------
        # Trade - Optional
        # ----------------------

        trade_value = convertir_bytes(record.VWGEW)

        if isinstance(trade_value, str):
            article.trade_value = trade_value

        # ----------------------
        # Unit - Optional
        # ----------------------

        unit_value = convertir_bytes(record.VWDIM)

        if not isinstance(unit_value, str):
            unit_value = ""

        article.unit_value = unit_value

        # ----------------------
        # Price - Optional
        # ----------------------

        article.price = self.prices_dict.get(cid_index, "")

        # ----------------------
        # Formula - Optional
        # ----------------------

        formule_tps = convertir_bytes(record.VWATR)

        if not isinstance(formule_tps, str):
            formula = materiaux_dyn = formula_obj = quantity = ""
        else:
            formula, materiaux_dyn, formula_obj, quantity = self.allplan.convertir_formule_bdd(formule_tps)

        article.formula = formula
        article.materiaux_dyn = materiaux_dyn
        article.formula_obj = formula_obj
        article.quantity = quantity

        # ----------------------

        self.cid_dict[cid_index] = article

        article_parent.children_article_list.append(article)

    def bcm_load_hierarchy(self, article_parent: BcmArticleMaterial, qs_parent: QStandardItem):

        if not isinstance(article_parent, BcmArticleMaterial) or not isinstance(qs_parent, QStandardItem):
            print("convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                  "not isinstance(article_parent, BcmArticle)")
            self.errors_list.append(f"bcm_load_hierarchy --  not isinstance(article_parent, BcmArticle)")
            return

        children_article_list = article_parent.children_article_list

        parent_type = qs_parent.data(user_data_type)

        for article in children_article_list:

            if not isinstance(article, BcmArticleMaterial):
                print("convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                      "not isinstance(article, BcmArticle)")
                self.errors_list.append(f"bcm_load_hierarchy --  not isinstance(article, BcmArticleMaterial)")
                continue

            # --------------
            # Folder creation
            # --------------

            if article.type_ele == folder_code:

                if parent_type == folder_code:

                    qs_folder_list = self.creation.folder_line(value=article.code,
                                                               description=article.desc,
                                                               tooltips=False)

                    if len(qs_folder_list) != col_cat_count:
                        print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                              f"len(qs_folder_list) != col_cat_count")
                        self.errors_list.append(f"bcm_load_hierarchy --  len(qs_folder_list) != col_cat_count")
                        continue

                    qs_folder: QStandardItem = qs_folder_list[col_cat_value]

                    if len(article.children_article_list) != 0:
                        self.bcm_load_hierarchy(article_parent=article, qs_parent=qs_folder)

                    qs_parent.appendRow(qs_folder_list)
                    continue

                print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                      f"The Folder: {article.code} can't be in a {parent_type} : {qs_parent.text()}")

                self.errors_list.append(f"bcm_load_hierarchy -- "
                                        f"The Folder: {article.code} can't be in a {parent_type} : {qs_parent.text()}")

                continue

            # --------------
            # Material creation
            # --------------

            if article.type_ele == material_code:

                if parent_type == folder_code:

                    if article.code == "":
                        article.code = find_new_title(base_title="No code", titles_list=self.material_upper_list)

                    elif article.code.upper() in self.material_upper_list:
                        article.code = find_new_title(base_title=article.code, titles_list=self.material_upper_list)

                    self.material_list.append(article.code)
                    self.material_upper_list.append(article.code.upper())

                    # -----------

                    qs_material_list = self.creation.material_line(value=article.code,
                                                                   description=article.desc,
                                                                   used_by_links=0,
                                                                   tooltips=False)

                    if len(qs_material_list) != col_cat_count:
                        print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                              f"len(qs_material_list) != col_cat_count")

                        self.errors_list.append(f"bcm_load_hierarchy -- len(qs_material_list) != col_cat_count")
                        continue

                    qs_material = qs_material_list[col_cat_value]

                    self.attributes_add(qs_value=qs_material, article=article)

                    if len(article.children_article_list) != 0:
                        self.bcm_load_hierarchy(article_parent=article, qs_parent=qs_material)

                    qs_parent.appendRow(qs_material_list)
                    continue

                print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                      f"The Material: {article.code} can't be in a {parent_type} : {qs_parent.text()}")

                self.errors_list.append(f"bcm_load_hierarchy -- The Material: {article.code} can't be in "
                                        f"a {parent_type} : {qs_parent.text()}")

                continue

            # --------------
            # Component creation
            # --------------

            if article.type_ele == component_code:

                if qs_parent.data(user_data_type) == material_code:

                    qs_component_list = self.creation.component_line(value=article.code,
                                                                     description=article.desc,
                                                                     tooltips=False)

                    if len(qs_component_list) != col_cat_count:
                        print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                              f"len(qs_component_list) != col_cat_count")

                        self.errors_list.append(f"bcm_load_hierarchy -- len(qs_component_list) != col_cat_count")
                        continue

                    qs_component = qs_component_list[col_cat_value]

                    self.attributes_add(qs_value=qs_component, article=article)

                    qs_parent.appendRow(qs_component_list)
                    continue

                print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                      f"The Component: {article.code} can't be in a {parent_type} : {qs_parent.text()}")

                self.errors_list.append(f"bcm_load_hierarchy -- The Component: {article.code} can't be in a"
                                        f" {parent_type} : {qs_parent.text()}")

                continue

            # --------------
            # Link creation
            # --------------

            if article.type_ele == link_code:

                if qs_parent.data(user_data_type) == material_code:

                    qs_link_list = self.creation.link_line(value=article.code,
                                                           description=article.desc,
                                                           tooltips=False)

                    if len(qs_link_list) != col_cat_count:
                        print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                              f"len(qs_link_list) != col_cat_count")
                        self.errors_list.append(f"bcm_load_hierarchy -- len(qs_component_list) != col_cat_count")
                        continue

                    self.link_list.append(article.code)

                    qs_parent.appendRow(qs_link_list)

                    material_name = qs_parent.text()

                    if not isinstance(material_name, str):
                        print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                              f"not isinstance(material_name, str)")
                        self.errors_list.append(f"bcm_load_hierarchy -- not isinstance(material_name, str)")
                        continue

                    self.material_with_link_list.append(material_name.upper())
                    continue

                print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                      f"The Link: {article.code} can't be in a {parent_type} : {qs_parent.text()}")

                self.errors_list.append(f"bcm_load_hierarchy -- The Link: {article.code} can't be in a "
                                        f"{parent_type} : {qs_parent.text()}")

                continue

            print(f"convert_manage -- ConvertBCmMaterial -- bcm_load_hierarchy -- "
                  f"The {article.type_ele}: {article.code} can't be in a {parent_type} : {qs_parent.text()}")

            self.errors_list.append(f"bcm_load_hierarchy -- The {article.type_ele}: {article.code} can't be in a"
                                    f" {parent_type} : {qs_parent.text()}")

    def verification(self, qs: QStandardItem) -> bool:

        children_count = qs.rowCount()

        if children_count == 0:
            return True

        type_ele_list = set()

        parent_type = qs.data(user_data_type)

        test = True

        for index_row in range(children_count):

            qs_child = qs.child(index_row, 0)

            type_ele = qs_child.data(user_data_type)

            if type_ele == link_code:
                type_ele = component_code

            if type_ele == attribute_code:
                continue

            if len(type_ele_list) == 0:

                type_ele_list.add(type_ele)

                if type_ele == folder_code and parent_type != folder_code:

                    print(f"convert_manage -- ConvertBCmMaterial -- verification -- "
                          f"error folder == {parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    self.errors_list.append("verification -- error folder == "
                                            f"{parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    test = False

                elif type_ele == material_code and parent_type != folder_code:

                    print(f"convert_manage -- ConvertBCmMaterial -- verification -- "
                          f"error parent material == {parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    self.errors_list.append("verification -- error material == "
                                            f"{parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    test = False

                elif type_ele == component_code and parent_type != material_code:

                    print(f"convert_manage -- ConvertBCmMaterial -- verification -- "
                          f"error parent component == {parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    self.errors_list.append("verification -- error component == "
                                            f"{parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    test = False

                elif type_ele == link_code and parent_type != material_code:

                    print(f"convert_manage -- ConvertBCmMaterial -- verification -- "
                          f"error parent link == {parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    self.errors_list.append("verification -- error link == "
                                            f"{parent_type}/{type_ele} -- {qs.text()}/{qs_child.text()}")

                    test = False

                    if qs_child.text().upper() not in self.material_article_dict:
                        print(f"convert_manage -- ConvertBCmMaterial -- verification -- "
                              f"error link == Material {qs_child.text()} doesn't exist")

                        self.errors_list.append("verification -- error link == "
                                                f"Material {qs_child.text()} doesn't exist")

                        test = False

                sub_test = self.verification(qs=qs_child)

                if not sub_test and test:
                    test = False

                continue

            if type_ele not in type_ele_list:
                print("convert_manage -- ConvertBCmMaterial -- verification -- error == type_ele not in type_ele_list :"
                      f"{qs.text()}/{qs_child.text()}")

                self.errors_list.append("verification -- error == type_ele not in type_ele_list== "
                                        f"Material {qs_child.text()} doesn't exist")

                test = False

                sub_test = self.verification(qs=qs_child)

                if not sub_test and test:
                    test = False

                continue

            sub_test = self.verification(qs=qs_child)

            if not sub_test and test:
                test = False

        return test

    @staticmethod
    def a___________________bcm_attribute__________________():
        pass

    def attributes_add(self, qs_value: QStandardItem, article: BcmArticleMaterial):

        # ---------------
        # Unit @202@
        # ---------------

        if article.unit_value != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.unit_value, number_str="202"))

        # ---------------
        # Price @203@
        # ---------------

        if article.price != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.price, number_str="203"))

        # ---------------
        # Full text @208@
        # ---------------

        if article.full_text != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.full_text, number_str="208"))

        # ---------------
        # Trade @209@
        # ---------------

        if article.trade_value != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.trade_value, number_str="209"))

        # ---------------

        if isinstance(qs_value, Material):
            return

        # ---------------
        # Object_filter @76@
        # ---------------

        if article.formula_obj != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.formula_obj, number_str="76"))

        # ---------------
        # Material_dyn @96@
        # ---------------

        if article.materiaux_dyn != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.materiaux_dyn, number_str="96"))

        # ---------------
        # Piece @215@
        # ---------------

        if article.quantity != "":
            qs_value.appendRow(self.creation.attribute_line(value=article.quantity, number_str="215"))

        # ---------------
        # Formula @267@
        # ---------------

        qs_value.appendRow(self.creation.attribute_line(value=article.formula, number_str="267"))

    @staticmethod
    def a___________________bcm_link__________________():
        pass

    def link_verification(self):

        links_set = set(self.link_list)

        link_errors = set()

        for material_name in links_set:

            if material_name in link_errors:
                continue

            if material_name not in self.material_list:
                self.link_delete_orphan(material_name=material_name)
                continue

            if material_name.upper() not in self.material_with_link_list:
                continue

            if not self.link_detect_circular(material_name=material_name, visited=set(), path=list()):
                continue

            link_errors.add(material_name)

    def link_detect_circular(self, material_name: str, visited: set, path: list) -> bool:

        if material_name in visited:
            loop_start_index = path.index(material_name)
            loop_path = " -> ".join(path[loop_start_index:])

            print(f"convert_manage -- ConvertBcmMaterial -- link_detect_circular -- loop detected "
                  f"({material_name})")

            loop_message = self.tr("Liens : Boucle détéctée")
            self.errors_list.append(f"{loop_message} : {loop_path} -> {material_name}")
            return True

        visited.add(material_name)
        path.append(material_name)

        search_start = self.cat_model.index(0, col_cat_value)

        search = self.cat_model.match(search_start, Qt.DisplayRole, material_name, -1,
                                      Qt.MatchExactly | Qt.MatchRecursive)

        for qm in search:

            if not qm_check(qm):
                print(f"convert_manage -- ConvertBcmMaterial -- link_detect_circular -- not qm_check(qm)")
                self.errors_list.append(f"link_detect_circular -- not qm_check(qm)")
                continue

            if qm.data(user_data_type) != link_code:
                continue

            qm_parent = qm.parent()

            if not qm_check(qm_parent):
                print(f"convert_manage -- ConvertBcmMaterial -- link_detect_circular -- not qm_check(qm_parent)")
                self.errors_list.append(f"link_detect_circular -- not qm_check(qm_parent)")
                continue

            parent_name = qm_parent.data()

            if not isinstance(parent_name, str):
                print(f"convert_manage -- ConvertBcmMaterial -- link_detect_circular -- "
                      f"not isinstance(parent_name, str)")
                self.errors_list.append(f"link_detect_circular -- not isinstance(parent_name, str)")
                continue

            if self.link_detect_circular(material_name=parent_name, visited=visited, path=path):
                return True

        visited.remove(material_name)
        path.pop()
        return False

    def link_delete_orphan(self, material_name: str):

        search_start = self.cat_model.index(0, col_cat_value)

        search = self.cat_model.match(search_start, Qt.DisplayRole, material_name, -1,
                                      Qt.MatchExactly | Qt.MatchRecursive)

        for qm in search:

            if not qm_check(qm):
                print(f"convert_manage -- ConvertBcmMaterial -- link_delete_orphan -- not isinstance(parent_name, str)")
                self.errors_list.append(f"link_delete_orphan -- not qm_check(qm)")
                continue

            if qm.data(user_data_type) != link_code:
                return False

            qm_parent = qm.parent()

            if not qm_check(qm_parent):
                print(f"convert_manage -- ConvertBcmMaterial -- link_delete_orphan -- not qm_check(qm_parent)")
                self.errors_list.append(f"link_delete_orphan -- not qm_check(qm_parent)")
                continue

            print(f"convert_manage -- ConvertBcmMaterial -- link_delete_orphan -- remove orphan link "
                  f"({qm.data()} in {qm_parent.data()})")

            self.errors_list.append(f"link_delete_orphan -- remove orphan link ({qm.data()} in {qm_parent.data()})")

            self.cat_model.removeRow(qm.row(), qm_parent)

        return True

    @staticmethod
    def a___________________bdd_bcm_price__________________():
        pass

    def get_all_prices(self):

        file_name = find_filename(file_path=self.file_path)

        file_name_new = file_name.lower().replace("vw", "pr")

        folder_path = find_folder_path(file_path=self.file_path)

        file_path = f"{folder_path}{file_name_new}.dbf"

        if not os.path.exists(file_path):
            return

        try:
            with dbf.Table(filename=file_path) as table:

                for record in table:

                    if dbf.is_deleted(record):
                        continue

                    price_float = convertir_bytes(record.PREPB)

                    if not isinstance(price_float, float):
                        continue

                    if price_float == 0:
                        continue

                    price_str = f"{price_float:.2f}"

                    current_code = convertir_bytes(record.PRIDPOS)

                    if not isinstance(current_code, str):
                        continue

                    if current_code in self.prices_dict:
                        continue

                    self.prices_dict[current_code] = price_str

        except Exception as error:
            print(f"convert_manage -- ConvertBcmComposants -- get_all_prices -- Erreur : {error}")
            self.errors_list.append(f"get_all_prices -- {error}")
            return

    @staticmethod
    def a___________________end__________________():
        pass


def a___________________favorites______():
    pass


class ConvertFavorite(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        pass

    # todo créer boite de dialogue de choix des attributs à importer

    def run(self):

        self.start_loading()

        try:

            parser = etree.XMLParser(recover=True)
            tree = etree.parse(self.file_path, parser=parser)
            root = tree.getroot()

            titre = self.tr("Favoris Allplan")

            dict_favoris_allplan = get_favorites_allplan_dict()

            for extension, nom_favoris in dict_favoris_allplan.items():

                if self.file_path.endswith(extension):
                    titre += f" {nom_favoris}"
                    break

            list_refus = ['4', '10', '94',
                          '101', '102', '103', '104', '105', '106', '108', '109', '112', '113', '114', '115',
                          '116', '117', '119', '133', '155', '156', '157', '161', '169', '171', '177',
                          '210', '253', '254', '255', '267',
                          '331', '333', '334', '336', '337', '338', '348', '373', '379', '383',
                          "689",
                          '748', '749',
                          '979',
                          '1414', '1442',
                          '1958', '1959', "1960", "1978", "1979",
                          '2100', '2101', '2102', '2103', '2105', '2105', '2110', '2111', '2112', '2113', '2114',
                          '2160',
                          '4302', '4303']

            liste_datas = list()

            qs_folder_list = self.creation.folder_line(value=titre, tooltips=False)
            qs_folder = qs_folder_list[col_cat_value]

            for attrib_set in root.iter("NEM_ATTRIB_SET"):

                dict_attributs = dict()

                for elem in attrib_set:

                    ifnr = elem.find("IFNR")

                    if not get_value_is_valid(ifnr):
                        continue

                    ifnr_text = ifnr.text

                    if ifnr_text == "":
                        continue

                    if ifnr_text in list_refus:
                        continue

                    value = elem.find("VALUE")

                    if value is None:
                        continue

                    value_text = value.text

                    if value_text is None:
                        value_text = ""

                    dict_attributs[ifnr_text] = value_text

                liste_datas.append(dict_attributs)

        except Exception as error:
            print(f"convert_manage -- ConvertFavorite -- run -- error : {error}")
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        for dict_attributs in liste_datas:

            dict_attributs: dict

            # -----------------------------------
            # Code
            # -----------------------------------

            if attribut_default_obj.current in dict_attributs:
                code: str = dict_attributs[attribut_default_obj.current]

                if code.strip() == "":

                    if "508" in dict_attributs:
                        code: str = dict_attributs["508"]
                    else:
                        continue
            else:
                continue

            code = code.strip()

            if code == "":
                continue

            # -----------------------------------
            # Description
            # -----------------------------------

            description = dict_attributs.get("207", "")

            # -----------------------------------
            # Material Création
            # -----------------------------------

            qs_material_list = self.creation.material_line(value=code, description=description, tooltips=False)

            qs_folder.appendRow(qs_material_list)

            qs_material: MyQstandardItem = qs_material_list[col_cat_value]

            liste_actuelle = list()

            # -----------------------------------
            # Attributes creation
            # -----------------------------------

            for number, value in dict_attributs.items():

                if number in liste_actuelle:
                    continue

                # -----------------------------------
                # Layer
                # -----------------------------------

                if number in attribute_val_default_layer:

                    if number != attribute_val_default_layer_first:
                        continue

                    for layer_number in attribute_val_default_layer:

                        attribute_obj = self.allplan.attributes_dict.get(layer_number)

                        if not isinstance(attribute_obj, AttributeDatas):
                            print(
                                "convert_manage -- ConvertFavorite -- run -- "
                                "not isinstance(attribute_obj, AttributeDatas) 1")
                            continue

                        if layer_number in dict_attributs:

                            value = dict_attributs.get(layer_number, "")

                        else:

                            value = attribute_val_default_layer.get(layer_number, "")

                        liste_actuelle.append(layer_number)

                        qs_attribute_list = self.creation.attribute_line(value=value,
                                                                         number_str=layer_number,
                                                                         model_enumeration=attribute_obj.enumeration)

                        index_insert = qs_material.get_attribute_insertion_index(number=layer_number)
                        qs_material.insertRow(index_insert, qs_attribute_list)

                    continue

                # -----------------------------------
                #  Fill
                # -----------------------------------

                if number in attribute_val_default_fill:

                    if number != attribute_val_default_fill_first:
                        continue

                    for fill_number in attribute_val_default_fill:

                        # -----------

                        if fill_number in dict_attributs:

                            value = dict_attributs.get(fill_number, "")

                        else:

                            value = attribute_val_default_fill.get(fill_number, "")

                        # -----------

                        if fill_number == "111":

                            id_hachurage = dict_attributs.get("118", attribute_val_default_fill.get("118", "0"))

                            if id_hachurage == "1":
                                fill_model = self.allplan.model_haching

                            elif id_hachurage == "2":
                                fill_model = self.allplan.model_pattern

                            elif id_hachurage == "3":
                                fill_model = self.allplan.model_color

                            elif id_hachurage == "5":
                                fill_model = self.allplan.model_style

                            else:
                                fill_model = None

                        else:

                            attribute_obj = self.allplan.attributes_dict.get(fill_number)

                            if not isinstance(attribute_obj, AttributeDatas):
                                print(
                                    "convert_manage -- ConvertFavorite -- run -- "
                                    "not isinstance(attribute_obj, AttributeDatas) 2")
                                continue

                            fill_model = attribute_obj.enumeration

                        liste_actuelle.append(fill_number)

                        qs_attribute_list = self.creation.attribute_line(value=value,
                                                                         number_str=fill_number,
                                                                         model_enumeration=fill_model)

                        index_insert = qs_material.get_attribute_insertion_index(number=fill_number)
                        qs_material.insertRow(index_insert, qs_attribute_list)

                    continue

                # -----------------------------------
                #  ROOM
                # -----------------------------------

                if number in attribute_val_default_room:

                    if number != attribute_val_default_room_first:
                        continue

                    for room_number in attribute_val_default_room:

                        attribute_obj = self.allplan.attributes_dict.get(room_number)

                        if not isinstance(attribute_obj, AttributeDatas):
                            print(
                                "convert_manage -- ConvertFavorite -- run -- "
                                "not isinstance(attribute_obj, AttributeDatas) 3")
                            continue

                        if room_number in dict_attributs:

                            value = dict_attributs.get(room_number, "")

                        else:

                            value = attribute_val_default_room.get(room_number, "")

                        liste_actuelle.append(room_number)

                        qs_attribute_list = self.creation.attribute_line(value=value,
                                                                         number_str=room_number,
                                                                         model_enumeration=attribute_obj.enumeration)

                        index_insert = qs_material.get_attribute_insertion_index(number=room_number)
                        qs_material.insertRow(index_insert, qs_attribute_list)

                    continue

                # -----------------------------------
                #  OTHER
                # -----------------------------------

                attribute_obj = self.allplan.attributes_dict.get(number)

                if not isinstance(attribute_obj, AttributeDatas):
                    print("convert_manage -- ConvertFavorite -- run -- not isinstance(attribute_obj, AttributeDatas) 4")
                    continue

                # -----------------------------------

                liste_actuelle.append(number)

                qs_attribute_list = self.creation.attribute_line(value=value,
                                                                 number_str=number,
                                                                 model_enumeration=attribute_obj.enumeration)

                index_insert = qs_material.get_attribute_insertion_index(number=number)
                qs_material.insertRow(index_insert, qs_attribute_list)

        self.cat_model.appendRow(qs_folder_list)
        self.end_loading()
        return True


def a___________________kukat______():
    pass


class ConvertKukat(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.folder_path = find_folder_path(file_path=file_path)

        self.chapitre = "chapters"
        self.extension = ".KAT"

        self.liste_code = list()
        self.datas = dict()

    def run(self):

        self.start_loading()

        # ---------

        first_file = f"{self.folder_path}{self.chapitre}{self.extension}"

        if not os.path.exists(first_file):
            print("convert_manage -- ConvertKuKat -- run -- not os.path.exists({first_file})")
            self.errors_list.append(f"run -- not os.path.exists({first_file})")
            self.end_loading()
            return False

        self.datas = self.read_kat_file(file_path=first_file)

        if len(self.datas) == 0:
            print("convert_manage -- ConvertKuKat -- run -- len(self.datas) == 0")
            self.errors_list.append(f"run -- len(self.datas) == 0")
            self.end_loading()
            return False

        self.read_children(self.datas)

        self.convert_to_model(self.datas, self.cat_model.invisibleRootItem())

        # ---------

        self.end_loading()

    def read_kat_file(self, file_path: str, ouvrage=False) -> dict:

        if not os.path.exists(file_path):
            return dict()

        lines_list = read_file_to_list(file_path=file_path)

        resultats = dict()

        if ouvrage:
            max_split = 2
        else:
            max_split = 1

        try:
            for ligne in lines_list[3:]:

                partie = ligne.strip().split(None, max_split)

                if len(partie) == 2 and not ouvrage:

                    reference, description = partie
                    resultats[reference] = {"description": description, "enfants": {}}

                elif len(partie) == 3 and ouvrage:

                    reference, unite, description = partie
                    resultats[reference] = {"code": reference, "description": description, "unite": unite, }

            return resultats

        except Exception as error:
            print(f"convert_manage -- ConvertKuKat -- read_kat_file -- error : {error}")
            self.errors_list.append(f"read_kat_file -- {error}")
            return dict()

    def read_children(self, data: dict):

        for key in data:

            fichier_actuel = f"{self.folder_path}{key}{self.extension}"

            if not os.path.exists(fichier_actuel):
                continue

            if "enfants" not in data[key]:
                return

            if len(key) == 3:
                data[key]["ouvrage"] = self.read_kat_file(fichier_actuel, ouvrage=True)

            else:
                data[key]["enfants"] = self.read_kat_file(fichier_actuel)
                self.read_children(data[key]["enfants"])

    def convert_to_model(self, data: dict, qs_dossier_parent: QStandardItem):

        for key, enfants in data.items():

            enfants: dict

            description = enfants.get("description", "")

            # -----------
            # Dossier
            # -----------

            if "enfants" in enfants:

                liste_qs_dossier_enfant = self.creation.folder_line(value=key,
                                                                    description=description,
                                                                    tooltips=self.conversion)

                qs_dossier_enfant = liste_qs_dossier_enfant[0]

                data_enfants = enfants["enfants"]

                # -----------
                # Ouvrage
                # -----------

                if "ouvrage" in enfants:
                    self.material_add(enfants["ouvrage"], qs_dossier_enfant)

                if len(data_enfants) != 0:
                    self.convert_to_model(data_enfants, qs_dossier_enfant)

                qs_dossier_parent.appendRow(liste_qs_dossier_enfant)

    def material_add(self, data: dict, qs_dossier_parent: MyQstandardItem):

        for key, enfants in data.items():

            enfants: dict

            if key in self.liste_code:
                return

            description = enfants.get("description", "")
            unite = enfants.get("unite", "")

            qs_list = self.creation.material_line(value=key, description=description)

            qs_material: MyQstandardItem = qs_list[0]

            if unite != "":
                qs_material.appendRow(self.creation.attribute_line(value=unite, number_str="202"))

            qs_dossier_parent.appendRow(qs_list)

            self.liste_code.append(key)


def a___________________extern______():
    pass


class ConvertExtern(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.datas = dict()

        self.root = None

        self.conversion_code = {"description": "207",
                                "unité": "202",
                                "prix": "203",
                                "description_contrat": "501",
                                "option": "502",
                                "variant": "503",
                                "niveau": "504"}

    def run(self):

        self.start_loading()

        self.root = xml_load_root(file_path=self.file_path)

        if not isinstance(self.root, etree._Element):
            print(f"convert_manage -- ConvertExtern -- run -- not isinstance(self.root, etree._Element)")
            self.errors_list.append(f"run -- not isinstance(self.root, etree._Element)")
            self.end_loading()
            return False

        # -------------------------------------------------
        # chargement hierarchie
        # -------------------------------------------------

        search_version = self.root.find("Dossier")

        try:

            if search_version is not None:
                self.catalog_load_old(self.cat_model.invisibleRootItem(), self.root)
            else:
                self.catalog_load(self.cat_model.invisibleRootItem(), self.root)

        except Exception as error:
            print(f"catalog_manage -- CatalogLoad --  analyse_display -- {error}")
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        self.end_loading()
        return True

    def catalog_load(self, qs_parent: QStandardItem, element: etree._Element):

        for child in element:

            tag = child.tag

            if not isinstance(tag, str):
                print("convert_manage -- ConvertExtern -- catalog_load -- not isinstance(tag, str)")
                self.errors_list.append(f"catalog_load -- not isinstance(tag, str)")
                continue

            tag = tag.capitalize()

            if tag == folder_code or tag == material_code or tag == component_code:

                if not self.verif_possibility(id_parent=id(qs_parent), type_ele=tag):
                    print("convert_manage -- ConvertExtern -- catalog_load -- not self.verif_possibility")
                    self.errors_list.append(f"catalog_load -- not self.verif_possibility")
                    continue

            # -----------------------------------------------
            # Node
            # -----------------------------------------------

            if tag == folder_code:

                name = child.get('name')

                if name is None:
                    print("convert_manage -- ConvertExtern -- catalog_load -- name is None")
                    self.errors_list.append(f"catalog_load -- name is None")
                    continue

                description = child.get('description', "")

                if "\n" in description:
                    description = description.replace("\n", "")

                qs_list = self.creation.folder_line(value=name,
                                                    description=description,
                                                    tooltips=False)

                qs_current = qs_list[0]

                self.catalog_load(qs_current, child)

                qs_parent.appendRow(qs_list)

                continue

            # -----------------------------------------------
            # Group and Position
            # -----------------------------------------------

            if tag == material_code or tag == component_code:
                self.children_load(child, qs_parent, tag)
                continue

    def children_load(self, child, qs_parent: QStandardItem, tag: str):

        presence_layer = False
        presence_remplissage = False
        presence_piece = False

        liste_defaut = list()
        datas_attribut_layer = dict(attribute_val_default_layer)
        datas_attribut_remp = dict(attribute_val_default_fill)
        datas_attribut_piece = dict(attribute_val_default_room)

        liste_autres = list()

        attributes_list = list()

        name = child.get("name", "")
        description = child.get("description", "")
        unit = child.get("unit", "")

        if unit != "":
            liste_defaut.append(["202", self.allplan.convert_unit(unit=unit)])

        # ----------------------------
        # search attributes
        # ----------------------------

        attributes = child.findall('Attribute')

        if len(attributes) != 0:

            for attribute in attributes:

                number_str = attribute.get("id")
                value = attribute.get("value", "")

                if number_str is None:
                    print("convert_manage -- ConvertExtern -- children_load -- number is None")
                    self.errors_list.append(f"children_load -- number_str is None")
                    return None

                if number_str in attributes_list:
                    print("convert_manage -- ConvertExtern -- children_load -- number in attributes_list")
                    self.errors_list.append(f"children_load -- number_str in attributes_list")
                    continue

                attributes_list.append(number_str)

                # -----------------------------------------
                # Attribute 83
                # -----------------------------------------

                if number_str == attribut_default_obj.current or number_str == "207" or number_str == "202":
                    continue

                # -----------------------------------------
                # Attribute 120  - 209 - 110
                # -----------------------------------------

                if number_str in attribut_val_defaut_defaut:
                    liste_defaut.append([number_str, value])
                    continue

                # -----------------------------------------
                # Attribute 141 - 349 - 346 - 345 - 347
                # -----------------------------------------

                if number_str in attribute_val_default_layer:
                    presence_layer = True
                    datas_attribut_layer[number_str] = value
                    continue

                # -----------------------------------------
                # Attribute 118 - 111 - 252 - 336 - 600
                # -----------------------------------------

                if number_str in attribute_val_default_fill:
                    presence_remplissage = True
                    datas_attribut_remp[number_str] = value
                    continue

                # -----------------------------------------
                # Attribute 231 - 235 - 232 - 266 - 233 - 264
                # -----------------------------------------

                if number_str in attribute_val_default_room:
                    presence_piece = True

                    if number_str == "232":
                        datas_attribut_piece[number_str] = self.allplan.traduire_valeur_232(value_current=value)

                    elif number_str == "233":
                        datas_attribut_piece[number_str] = self.allplan.traduire_valeur_233(value_current=value)

                    elif number_str == "235":
                        datas_attribut_piece[number_str] = self.allplan.traduire_valeur_235(value_current=value)

                    else:
                        datas_attribut_piece[number_str] = value

                    continue

                # -----------------------------------------
                # Attribute 76 - 96 - 180 - 267
                # -----------------------------------------

                if number_str in formula_list_attributes:

                    if self.allplan.version_allplan_current != "2022":

                        if len(re.findall(pattern=formula_piece_pattern, string=value)):
                            value = re.sub(pattern=formula_piece_pattern,
                                           repl=lambda m: formula_piece_dict.get(m.group(0)),
                                           string=value,
                                           flags=re.IGNORECASE)

                    if number_str == "267" and value == "":
                        value = self.allplan.recherche_formule_defaut(unit=unit)

                    liste_autres.append([number_str, value])
                    continue

                # -----------------------------------------
                # Attribute > 1999 & < 12000
                # -----------------------------------------

                attribute_obj = self.allplan.attributes_dict.get(number_str)

                if not isinstance(attribute_obj, AttributeDatas):
                    print("convert_manage -- ConvertExtern -- children_load -- "
                          "not isinstance(attribute_obj, AttributeDatas)")
                    self.errors_list.append(f"children_load -- not isinstance(attribute_obj, AttributeDatas)")
                    continue

                if attribute_obj.user:

                    if attribute_obj.option not in [code_attr_formule_str, code_attr_formule_int,
                                                    code_attr_formule_float]:
                        liste_autres.append([number_str, value])
                        continue

                    valeur_formule = attribute_obj.value

                    if self.allplan.version_allplan_current != "2022":
                        if len(re.findall(pattern=formula_piece_pattern, string=valeur_formule)):
                            valeur_formule = re.sub(pattern=formula_piece_pattern,
                                                    repl=lambda m: formula_piece_dict.get(m.group(0)),
                                                    string=valeur_formule,
                                                    flags=re.IGNORECASE)

                    liste_autres.append([number_str, valeur_formule])
                    continue

                # -----------------------------------------
                # Attribute Other
                # -----------------------------------------

                liste_autres.append([number_str, value])
                continue

            liste_defaut.sort(key=lambda x: int(x[0]))
            liste_autres.sort(key=lambda x: int(x[0]))

        # ==================================================================== #
        # QStandardItem Creation (Material / Component
        # ==================================================================== #

        if name == "":
            print("convert_manage -- ConvertExtern -- children_load -- name is empty")
            return None

        if tag == material_code:

            qs_list = self.creation.material_line(value=name,
                                                  description=description,
                                                  tooltips=False)

            qs_current: QStandardItem = qs_list[0]

        elif tag == component_code:

            qs_list = self.creation.component_line(value=name,
                                                   description=description,
                                                   tooltips=False)

            qs_current: QStandardItem = qs_list[0]

        else:
            print(f"convert_manage -- ConvertExtern -- children_load -- tag is wrong : {tag}")
            self.errors_list.append(f"children_load -- tag is wrong : {tag}")
            return None

        # ==================================================================== #
        # Attribute Creation
        # ==================================================================== #

        # -----------------------------------------
        # Attribute 120  - 209 - 110
        # -----------------------------------------

        for number_str, value in liste_defaut:
            qs_current.appendRow(self.creation.attribute_line(value=value, number_str=number_str))

        # -----------------------------------------
        # Attribute 118 - 111 - 252 - 336 - 600
        # -----------------------------------------

        if presence_remplissage:

            type_remplissage = datas_attribut_remp.get("118", "0")

            if type_remplissage == "1":
                model_enumeration = self.allplan.model_haching
            elif type_remplissage == "2":
                model_enumeration = self.allplan.model_pattern
            elif type_remplissage == "3":
                model_enumeration = self.allplan.model_color
            elif type_remplissage == "5":
                model_enumeration = self.allplan.model_style
            elif type_remplissage == "6":
                model_enumeration = self.allplan.model_none
            else:
                model_enumeration = self.allplan.model_none
                datas_attribut_remp["111"] = "-1"
                datas_attribut_remp["252"] = "-1"
                datas_attribut_remp["336"] = ""
                datas_attribut_remp["600"] = "0"

            for number_str, value in datas_attribut_remp.items():

                if number_str == "111":
                    qs_current.appendRow(self.creation.attribute_line(value=value,
                                                                      number_str=number_str,
                                                                      model_enumeration=model_enumeration))
                    continue

                qs_current.appendRow(self.creation.attribute_line(value=value, number_str=number_str))

        # -----------------------------------------
        # Attribute 141 - 349 - 346 - 345 - 347
        # -----------------------------------------

        if presence_layer:
            for number_str, value in datas_attribut_layer.items():
                qs_current.appendRow(self.creation.attribute_line(value=value, number_str=number_str))

        # -----------------------------------------
        # Attribute 231 - 235 - 232 - 266 - 233 - 264
        # -----------------------------------------

        if presence_piece:

            for number_str, value in datas_attribut_piece.items():
                qs_current.appendRow(self.creation.attribute_line(value=value, number_str=number_str))

        # -----------------------------------------
        # Attributes Other
        # -----------------------------------------

        if len(liste_autres) != 0:
            for number_str, value in liste_autres:
                qs_current.appendRow(self.creation.attribute_line(value=value, number_str=number_str))

        # -----------------------------------------

        if tag == material_code:
            self.catalog_load(qs_current, child)

        qs_parent.appendRow(qs_list)

    def verif_possibility(self, id_parent: int, type_ele: str) -> bool:

        if id_parent not in self.datas:
            self.datas[id_parent] = type_ele
            return True

        return type_ele == self.datas[id_parent]

    def catalog_load_old(self, qs_parent: QStandardItem, element: etree._Element):

        component_in = element.find("Composant") is not None
        folder_in = element.find("Dossier") is not None

        if folder_in and component_in:
            folder_creation = True
        else:
            folder_creation = False

        new_folder = None

        for child in element:

            tag = child.tag

            if not isinstance(tag, str):
                print("convert_manage -- ConvertExtern -- catalog_load_old -- not isinstance(tag, str)")
                self.errors_list.append(f"catalog_load_old -- not isinstance(tag, str)")
                continue

            if tag != "Dossier" and tag != "Composant":
                print(f"convert_manage -- ConvertExtern -- catalog_load_old -- tag is wrong : {tag}")
                self.errors_list.append(f"catalog_load_old -- wrong tag")
                continue

            # -----------------------------------------------
            # Folder
            # -----------------------------------------------

            if tag == "Dossier":

                name = child.get('name')

                if not isinstance(name, str):
                    print(f"convert_manage -- ConvertExtern -- catalog_load_old -- folder -- not isinstance(name, str)")
                    self.errors_list.append(f"catalog_load_old -- folder -- not isinstance(name, str)")
                    continue

                if " - " not in name:
                    description = ""
                else:
                    name, description = name.split(" - ", maxsplit=1)

                qs_list = self.creation.folder_line(value=name,
                                                    description=description,
                                                    tooltips=False)

                qs_current = qs_list[0]

                qs_parent.appendRow(qs_list)

                self.catalog_load_old(qs_parent=qs_current, element=child)

                continue

            # -----------------------------------------------
            # Component
            # -----------------------------------------------

            if tag == "Composant":

                name = child.get('name')

                if not isinstance(name, str):
                    print(f"convert_manage -- ConvertExtern -- catalog_load_old -- Composant -- "
                          f"not isinstance(name, str)")
                    self.errors_list.append(f"catalog_load_old -- component -- not isinstance(name, str)")
                    continue

                ele_description = child.find("description")

                if ele_description is None:
                    description = ""
                else:
                    description = ele_description.text

                    if "\n" in description:
                        description = description.replace("\n", " ")

                    if not isinstance(description, str):
                        description = ""

                qs_list = self.creation.component_line(value=name,
                                                       description=description,
                                                       tooltips=False)

                unit_ele = child.find("Unité")

                if unit_ele is None:
                    qs_parent.appendRow(qs_list)
                    continue

                unit = unit_ele.text

                if not isinstance(unit, str):
                    qs_parent.appendRow(qs_list)
                    continue

                qs_current = qs_list[0]

                qs_current.appendRow(self.creation.attribute_line(value=unit,
                                                                  number_str="202",
                                                                  model_enumeration=self.allplan.model_units))

                formula = self.allplan.recherche_formule_defaut(unit=unit)
                qs_current.appendRow(self.allplan.creation.attribute_line(value=formula, number_str="267"))

                if folder_creation:

                    if isinstance(new_folder, Folder):

                        new_folder.appendRow(qs_list)

                    else:

                        qs_folder_list = self.creation.folder_line(value=name,
                                                                   description=description,
                                                                   tooltips=False)

                        new_folder = qs_folder_list[0]

                        qs_parent.appendRow(qs_folder_list)

                        new_folder.appendRow(qs_list)

                else:

                    qs_parent.appendRow(qs_list)


def a___________________excel______():
    pass


class ObjExcel:

    def __init__(self):
        super().__init__()

        self.name = ""
        self.description = ""
        self.his_parent = None

        self.children_type = None

        self.children = list()
        self.attributes = dict()


class RootExcel(ObjExcel):

    def __init__(self):
        super().__init__()

    def add_children(self, children) -> bool:
        if not isinstance(children, FolderExcel):
            print(f"convertmanage -- ConvertExcel -- RootExcel -- can't add this children in root")
            return False

        self.children.append(children)
        return True

    @staticmethod
    def add_attribute(number: str, value: str):
        return


class FolderExcel(ObjExcel):

    def __init__(self, his_parent, name: str):
        super().__init__()

        self.name = name
        self.his_parent = his_parent

    def add_attribute(self, number: str, value: str):

        if number != "207":
            return

        if not isinstance(value, str):
            value = ""

        self.description = value

    def add_children(self, children) -> bool:

        if isinstance(children, FolderExcel):
            if self.children_type == folder_code:
                self.children.append(children)
                return True

            if self.children_type is None:
                self.children_type = folder_code
                self.children.append(children)
                return True

        if isinstance(children, MaterialExcel):
            if self.children_type == material_code:
                self.children.append(children)
                return True

            if self.children_type is None:
                self.children_type = material_code
                self.children.append(children)
                return True

            print("convertmanage -- ConvertExcel -- FolderExcel -- "
                  f"can't add {folder_code} -> {material_code} already exists,")

            return False

        if isinstance(children, ComponentExcel):
            if self.children_type == component_code:
                self.children.append(children)
                return True

            if self.children_type is None:
                self.children_type = component_code
                self.children.append(children)
                return True

            print("convertmanage -- ConvertExcel -- FolderExcel -- "
                  f"can't add {folder_code} -> {component_code} already exists,")
            return False

        print("convertmanage -- ConvertExcel -- FolderExcel -- can't add unknown type,")
        return False


class MaterialExcel(ObjExcel):

    def __init__(self, his_parent: FolderExcel, name: str):
        super().__init__()

        self.name = name
        self.his_parent = his_parent

    def add_attribute(self, number: str, value: str):

        if number == "207":

            if not isinstance(value, str):
                value = ""

            self.description = value

            return

        if number in self.attributes:
            return

        self.attributes[number] = value

    def add_children(self, children) -> bool:

        if not isinstance(children, ComponentExcel):
            print(f"convertmanage -- ConvertExcel -- MaterialExcel -- can't add this children in {material_code}")
            return False

        self.children.append(children)
        return True


class ComponentExcel(ObjExcel):

    def __init__(self, his_parent: Union[FolderExcel, MaterialExcel], name: str):
        super().__init__()

        self.name = name
        self.his_parent = his_parent

    def add_attribute(self, number: str, value: str):

        if number == "207":

            if not isinstance(value, str):
                value = ""

            self.description = value

            return

        if number in self.attributes:
            return

        self.attributes[number] = value

    @staticmethod
    def add_children(_) -> bool:
        print(f"convertmanage -- ConvertExcel -- Component -- can't add children in {component_code}")
        return False


class ConvertExcel(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.root = RootExcel()

        # --------------

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        self.start_loading()

        workbook = excel_load_workbook(file_path=self.file_path)

        if not isinstance(workbook, openpyxl.Workbook):
            self.errors_list.append("run -- not isinstance(workbook, openpyxl.Workbook)")
            self.end_loading()
            return False

        try:

            sheet = workbook.active

            colum_count = sheet.max_column

            if not isinstance(colum_count, int):
                # noinspection PyArgumentList
                sheet.calculate_dimension(force=True)

                colum_count = sheet.max_column

            if not isinstance(colum_count, int):
                print("convertmanage -- ConvertExcel -- run -- nb_col is None")
                self.errors_list.append("run -- not isinstance(colum_count, int)")
                self.end_loading()
                return False

            if colum_count == 0:
                print("convertmanage -- ConvertExcel -- run -- max_column_for_row == 0")
                self.errors_list.append("run -- colum_count == 0")
                self.end_loading()
                return False

            columns_dict = dict()

            for column_index in range(1, colum_count + 1):

                obj_title = sheet.cell(1, column_index)

                if obj_title is None:
                    print("convertmanage -- ConvertExcel -- run -- obj_titre is None")
                    self.errors_list.append("run -- obj_title is None")
                    self.end_loading()
                    return False

                value = obj_title.value

                if isinstance(value, int):
                    value = f"{value}"

                if not isinstance(value, str):
                    print("convertmanage -- ConvertExcel -- run -- not isinstance(valeur, str)")
                    self.errors_list.append("run -- not isinstance(value, str)")
                    self.end_loading()
                    return False

                value = value.strip().lower()

                if value == "":
                    print("convertmanage -- ConvertExcel -- run -- value empty")
                    break

                columns_dict[value] = column_index - 1

            if len(columns_dict) != 4:
                result = self.load_component_excel(sheet=sheet, columns_dict=columns_dict)

            elif 'type' in columns_dict and 'parent' in columns_dict and 'attribute' in columns_dict \
                    and 'value' in columns_dict:

                result = self.load_material_excel(sheet=sheet, columns_dict=columns_dict)
            else:

                result = self.load_component_excel(sheet=sheet, columns_dict=columns_dict)

            self.end_loading()

            return result

        except Exception as error:
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

    @staticmethod
    def a___________________material______():
        pass

    def load_material_excel(self, sheet, columns_dict: dict) -> bool:

        type_index = columns_dict["type"]
        parent_index = columns_dict["parent"]
        attribute_index = columns_dict["attribute"]
        value_index = columns_dict["value"]

        current_parent = self.root
        current_obj = None

        # -------------------------
        # Read lines
        # -------------------------

        for row in sheet.iter_rows(min_row=2, values_only=True):

            # ----------------------
            # read line
            # ----------------------

            if len(row) < 4:
                print("convertmanage -- ConvertExcel -- load_material_excel -- len(row) < 4")
                continue

            try:

                type_txt: str = row[type_index]
                parent_txt: str = row[parent_index]
                number_txt: str = row[attribute_index]
                value_txt: str = row[value_index]

            except Exception as error:
                print(f"convertmanage -- ConvertExcel -- load_material_excel -- error : {error}")
                self.errors_list.append(f"load_material_excel -- {error}")
                continue

            # ----------------------
            # find value
            # ----------------------

            if isinstance(value_txt, int):
                value_txt = f"{value_txt}"

            elif not isinstance(value_txt, str):
                value_txt = ""

            # ----------------------
            # find type element
            # ----------------------

            if not isinstance(type_txt, str):
                continue

            if type_txt == "":
                continue

            if type_txt.lower() in attribute_code_list:

                if not isinstance(current_obj, (FolderExcel, MaterialExcel, ComponentExcel)):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- attribute -- error object")
                    self.errors_list.append(f"load_material_excel -- error object")
                    continue

                # ----------------------
                # défintion du numéro
                # ----------------------

                if isinstance(number_txt, int):
                    number_txt = f"{number_txt}"

                elif isinstance(number_txt, str):

                    try:
                        number_txt = str(int(number_txt))

                    except Exception as error:
                        print("convertmanage -- ConvertExcel -- load_material_excel -- attribute -- "
                              "error convert number")
                        self.errors_list.append(f"load_material_excel -- {error}")
                        continue

                current_obj.add_attribute(number=number_txt, value=value_txt)
                continue

            # ----------------------
            # Search parent
            # ----------------------

            if isinstance(parent_txt, int):
                parent_txt = ""

            elif not isinstance(parent_txt, str):
                parent_txt = ""

            # Si parent est vide alors root
            if parent_txt == "":
                current_parent = self.root

            elif current_parent is None:
                current_parent = self.root

            elif current_parent.name == parent_txt:

                if isinstance(current_parent, MaterialExcel) and type_txt == material_code:
                    current_parent = current_parent.his_parent

            elif current_parent.name != parent_txt:
                current_parent = self.get_parent(obj_excel=current_parent, parent_name=parent_txt)

                if current_parent is None:
                    print("convertmanage -- ConvertExcel -- load_material_excel -- current_parent is None")
                    self.errors_list.append(f"load_material_excel -- current_parent is None")
                    continue

            # ----------------------
            # Folder
            # ----------------------

            if type_txt.lower() in folder_code_list:

                obj = FolderExcel(his_parent=current_parent, name=value_txt)

                if not current_parent.add_children(obj):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- folder -- error add_children(obj)")
                    self.errors_list.append(f"load_material_excel -- not current_parent.add_children(obj)")
                    continue

                current_parent = current_obj = obj

            # ----------------------
            # Material
            # ----------------------

            elif type_txt.lower() in material_code_list:

                obj = MaterialExcel(his_parent=current_parent, name=value_txt)

                if not current_parent.add_children(obj):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- material -- error add_children(obj)")
                    self.errors_list.append(f"load_material_excel -- not current_parent.add_children(obj) 2")
                    continue

                current_parent = current_obj = obj

            # ----------------------
            # Component
            # ----------------------

            elif type_txt.lower() in component_code_list:

                obj = ComponentExcel(his_parent=current_parent, name=value_txt)

                if not current_parent.add_children(obj):
                    print("convertmanage -- ConvertExcel -- load_material_excel -- component -- "
                          "error add_children(obj)")
                    self.errors_list.append(f"load_material_excel -- not current_parent.add_children(obj) 3")
                    continue

                current_obj = obj

        self.model_add(qs_parent=self.cat_model.invisibleRootItem(), obj_excel=self.root)

        return True

    def get_parent(self, obj_excel: Union[RootExcel, FolderExcel, MaterialExcel, ComponentExcel], parent_name: str):

        if obj_excel is None or obj_excel == self.root:
            print(f"convertmanage -- ConvertExcel -- get_parent -- obj_excel is None")
            self.errors_list.append(f"get_parent -- obj_excel is None")
            return None

        obj_parent = obj_excel.his_parent

        if not isinstance(obj_parent, (RootExcel, FolderExcel, MaterialExcel, ComponentExcel)):
            print(f"convertmanage -- ConvertExcel -- get_parent -- bad obj_parent")
            self.errors_list.append(f"get_parent -- bad obj_parent")
            return None

        if obj_parent.name == parent_name:
            return obj_parent

        children_list = obj_parent.children

        for obj_child in children_list:

            if not isinstance(obj_child, (FolderExcel, MaterialExcel, ComponentExcel)):
                print(f"convertmanage -- ConvertExcel -- get_parent -- bad obj_child")
                self.errors_list.append(f"get_parent -- bad obj_child")
                continue

            if obj_child.name == parent_name:
                return obj_child

        return self.get_parent(obj_excel=obj_parent, parent_name=parent_name)

    def model_add(self, qs_parent: QStandardItem,
                  obj_excel: Union[RootExcel, FolderExcel, MaterialExcel, ComponentExcel]):

        if not isinstance(obj_excel, (RootExcel, FolderExcel, MaterialExcel, ComponentExcel)):
            print(f"convertmanage -- ConvertExcel -- model_add -- bad object")
            self.errors_list.append(f"model_add -- bad object")
            return

        if len(obj_excel.attributes) != 0:

            for number, value in obj_excel.attributes.items():

                model_111 = None

                if number == "111":

                    value_118 = obj_excel.attributes.get("118", 0)

                    if value_118 == "1":
                        model_111 = self.allplan.model_haching
                    elif value_118 == "2":
                        model_111 = self.allplan.model_pattern
                    elif value_118 == "3":
                        model_111 = self.allplan.model_color
                    elif value_118 == "5":
                        model_111 = self.allplan.model_style

                qs_list = self.creation.attribute_line(value=value, number_str=number, model_enumeration=model_111)

                qs_parent.appendRow(qs_list)

        children = obj_excel.children

        if len(children) == 0:
            return

        for child in children:

            child: ObjExcel

            name = child.name
            description = child.description

            if isinstance(child, FolderExcel):
                qs_list = self.creation.folder_line(value=name, description=description, tooltips=False)

            elif isinstance(child, MaterialExcel):

                qs_list = self.creation.material_line(value=name, description=description, tooltips=False)

            elif isinstance(child, ComponentExcel):
                qs_list = self.creation.component_line(value=name, description=description, tooltips=False)

            else:
                print(f"convertmanage -- ConvertExcel -- model_add -- bad child")
                self.errors_list.append(f"model_add -- bad child")
                return

            self.model_add(qs_parent=qs_list[0], obj_excel=child)

            qs_parent.appendRow(qs_list)

    @staticmethod
    def a___________________component______():
        pass

    def load_component_excel(self, sheet, columns_dict: dict) -> bool:

        columns_base = dict()
        columns_other = dict()

        title = find_filename(self.file_path)

        if not isinstance(title, str):
            title = "Excel"

        qs_folder_list = self.creation.folder_line(value=title, tooltips=False)
        root: MyQstandardItem = qs_folder_list[0]

        for value, column_index in columns_dict.items():

            if not isinstance(value, str):
                print("convertmanage -- ConvertExcel -- load_component_excel -- not isinstance(valeur, str)")
                self.errors_list.append(f"load_component_excel -- not isinstance(valeur, str)")
                return False

            try:

                int(value)
                number = value

            except Exception:

                number = self.allplan.find_number_by_name(name=value)

                if number == "":
                    print("convertmanage -- ConvertExcel -- load_component_excel -- number not found")
                    self.errors_list.append(f"load_component_excel -- number not found")
                    continue

            if (number in attribute_val_default_layer or number in attribute_val_default_fill or
                    number in attribute_val_default_room):
                continue

            if number in columns_other:
                print("convertmanage -- ConvertExcel -- load_component_excel -- number in columns_dict")
                self.errors_list.append(f"load_component_excel -- number in columns_dict")
                continue

            if number == "83" or number == "207":
                columns_base[number] = column_index
            else:
                columns_other[number] = column_index

        if "83" not in columns_base:
            print("convertmanage -- ConvertExcel -- load_component_excel -- 83 not in columns_base")
            self.errors_list.append(f"load_component_excel -- 83 not in columns_base")
            return False

        description_in = "207" in columns_base

        # -------------------------
        # Sort header
        # -------------------------

        order_list_columns = {}

        for number in liste_attributs_ordre:
            if number in columns_other:
                order_list_columns[number] = columns_other[number]

        other_columns = [number for number in columns_other if number not in order_list_columns]

        try:

            other_columns.sort(key=int)

        except Exception:
            pass

        for number in other_columns:
            order_list_columns[number] = columns_other[number]

        # -------------------------
        # Read lines
        # -------------------------

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if not isinstance(row, tuple):
                print("convertmanage -- ConvertExcel -- load_component_excel -- not isinstance(row, tuple)")
                self.errors_list.append(f"load_component_excel -- not isinstance(row, tuple)")
                continue

            column_count = len(row) - 1

            code = row[columns_base["83"]]

            if description_in:
                description = row[columns_base["207"]]
            else:
                description = ""

            qs_component_list = self.creation.component_line(value=code, description=description, tooltips=False)

            qs_component: Component = qs_component_list[0]

            for number, column in order_list_columns.items():

                if column > column_count:
                    print("convertmanage -- ConvertExcel -- load_component_excel -- column > column_count")
                    self.errors_list.append(f"load_component_excel -- column > column_count")
                    continue

                value = row[column]

                if isinstance(value, int):
                    value = f"{value}"

                elif isinstance(value, float):
                    value = f"{value:.3f}"

                elif not isinstance(value, str):
                    value = ""

                qs_attribute_list = self.creation.attribute_line(value=value, number_str=number)

                qs_component.appendRow(qs_attribute_list)

            root.appendRow(qs_component_list)

        self.cat_model.appendRow(qs_folder_list)

        return True

    @staticmethod
    def a___________________end______():
        pass


def a___________________csv______():
    pass


class ConvertCSV(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.root = RootExcel()
        self.separator = "\t"
        self.bdd_type = False

        # --------------

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        self.bdd_type = msg(titre=application_title,
                            message=self.tr("Définir le type de base de données"),
                            bt_ok=self.tr("Ouvrage"),
                            bt_no=self.tr("Composant"),
                            type_bouton=QMessageBox.Ok | QMessageBox.No,
                            icone_question=True,
                            defaut_bouton=QMessageBox.Ok) == QMessageBox.Ok

        # -------------

        self.start_loading()

        # -------------
        try:

            datas = read_file_to_list(file_path=self.file_path)

            if not isinstance(datas, list):
                print(f"convert_manage -- ConvertCSV -- run -- not isinstance(datas, dict)")
                self.errors_list.append(f"run -- not isinstance(datas, list)")
                self.end_loading()
                return False

            if len(datas) < 2:
                print(f"convert_manage -- ConvertCSV -- run -- len(datas) < 2")
                self.errors_list.append(f"run -- len(datas) < 2")
                self.end_loading()
                return False

            header = datas[0]
            datas.pop(0)

            if not isinstance(header, str):
                print(f"convert_manage -- ConvertCSV -- run --  not isinstance(header, str)")
                self.errors_list.append(f"run -- not isinstance(header, str)")
                self.end_loading()
                return False

            header = header.strip()

            if self.separator not in header:

                find_separator = False

                for separator in [";", "|"]:
                    if separator in header:
                        self.separator = separator
                        find_separator = True
                        break

                if not find_separator:
                    print(f"convert_manage -- ConvertCSV -- run -- self.separator not in header")
                    self.errors_list.append(f"run -- not find_separator")
                    self.end_loading()
                    return False

            header_list = header.split(self.separator)

            columns_base = dict()
            columns_other = dict()

            for column_index, value in enumerate(header_list):

                if not isinstance(value, str):
                    print("convert_manage -- ConvertCSV -- run -- not isinstance(valeur, str)")
                    self.errors_list.append(f"run -- not isinstance(valeur, str)")
                    self.end_loading()
                    return False

                value = value.strip()

                if value == "0" or value == "":
                    continue

                try:

                    int(value)
                    number = value

                except Exception:

                    number = self.allplan.find_number_by_name(name=value)

                    if number == "":
                        continue

                if (number in attribute_val_default_layer or number in attribute_val_default_fill or
                        number in attribute_val_default_room):
                    continue

                if number in columns_other:
                    print("convert_manage -- ConvertCSV -- run -- number in columns_dict")
                    self.errors_list.append(f"run -- number in columns_dict")
                    continue

                if number == "83" or number == "207":
                    columns_base[number] = column_index
                else:
                    columns_other[number] = column_index

            if "83" not in columns_base:
                print("convert_manage -- ConvertCSV -- run -- 83 not in columns_base")
                self.errors_list.append(f"run -- 83 not in columns_base")
                self.end_loading()
                return False

            description_in = "207" in columns_base

            # -------------------------
            # Sort header
            # -------------------------

            order_list_columns = {}

            for number in liste_attributs_ordre:
                if number in columns_other:
                    order_list_columns[number] = columns_other[number]

            other_columns = [number for number in columns_other if number not in order_list_columns]

            other_columns.sort(key=int)

            for number in other_columns:
                order_list_columns[number] = columns_other[number]

            # -------------------------
            # Read lines
            # -------------------------

            title = find_filename(self.file_path)

            if not isinstance(title, str):
                title = "Excel"

            qs_folder_list = self.creation.folder_line(value=title, tooltips=False)
            root: MyQstandardItem = qs_folder_list[0]

            for line in datas:

                if not isinstance(line, str):
                    print("convert_manage -- ConvertCSV -- run -- not isinstance(line, str):")
                    self.errors_list.append(f"run -- not isinstance(line, str)")
                    continue

                line = line.strip()

                if self.separator not in line:
                    continue

                line_list = line.split(self.separator)

                column_count = len(line_list)

                code = line_list[columns_base["83"]]

                if description_in:
                    description = line_list[columns_base["207"]]
                else:
                    description = ""

                if self.bdd_type:

                    qs_component_list = self.creation.material_line(value=code, description=description, tooltips=False)

                else:

                    qs_component_list = self.creation.component_line(value=code, description=description,
                                                                     tooltips=False)

                qs_component: Component = qs_component_list[0]

                for number, column in order_list_columns.items():

                    if column > column_count:
                        print("convert_manage -- ConvertCSV -- run -- column > column_count")
                        self.errors_list.append(f"run -- column > column_count")
                        continue

                    value = line_list[column]

                    if not isinstance(value, str):
                        value = ""

                    qs_component.appendRow(self.creation.attribute_line(value=value, number_str=number))

                root.appendRow(qs_component_list)

            self.cat_model.appendRow(qs_folder_list)

        except Exception as error:
            print(f"convert_manage -- ConvertCSV -- run -- error : {error}")
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        self.end_loading()
        return True

    @staticmethod
    def a___________________end______():
        pass


def a___________________mxdb______():
    pass


class ConvertMXDB(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.folder_dict = dict()
        self.unit_dict = dict()

        # --------------

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        self.start_loading()

        # ---------------------------------
        # Connection DB
        # ---------------------------------

        try:
            conn = sqlite3.connect(self.file_path)

            cursor = conn.cursor()

            # ---------------------------------
            # Folder level 0 & Folder level 1
            # ---------------------------------

            self.get_folders(cursor=cursor)

            # ---------------------------------
            # Material & Component
            # ---------------------------------

            self.get_children(cursor=cursor)

            cursor.close()

        except Exception as error:

            print(f"convert_manage -- ConvertXpwe -- run -- error : {error}")

            # -----------------

            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        # -----------------

        self.end_loading()
        return True

    def get_folders(self, cursor):

        parent_id_col = 0
        code_col = 1
        desc_col = 2
        my_id_col = 3

        cursor.execute("""SELECT PUNTLIVART, CODLIVART, NOMELIVART1, ID_Uni FROM FLivelliArt ORDER BY CODLIVART""")

        rows = cursor.fetchall()

        # -----------------

        for row in rows:

            my_id = row[my_id_col]

            if not isinstance(my_id, int):
                print("convert_manage -- ConvertMXDB -- get_folders -- not isinstance(my_id, str)")
                self.errors_list.append("get_folders -- not isinstance(my_id, int)")
                continue

            if my_id == "":
                print("convert_manage -- ConvertMXDB -- get_folders -- my_id are empty")
                self.errors_list.append("get_folders -- my_id are empty")
                continue

            # ------- Code -------

            code = row[code_col]

            if not isinstance(code, str):
                print("convert_manage -- ConvertMXDB -- get_folders -- not isinstance(code, str)")
                self.errors_list.append("get_folders -- not isinstance(code, str)")
                continue

            code = code.replace(" ", "")
            code = code.strip()

            # ------- Description -------

            desc = row[desc_col]

            if not isinstance(desc, str):
                desc = ""
            else:
                desc = desc.strip()

            # ------- Check Valide -------

            if code == "":

                if desc.upper().startswith("ATTENZIONE:"):
                    continue

                print("convert_manage -- ConvertMXDB -- get_folders -- code are empty")
                self.errors_list.append(f"get_folders -- code are empty --> {my_id} : {desc}")
                continue

            # ------- Parent ID -------

            parent_id = row[parent_id_col]

            if not isinstance(parent_id, int):
                print("convert_manage -- ConvertMXDB -- get_folders -- not isinstance(parent_id, int)")
                self.errors_list.append("get_folders -- not isinstance(parent_id, int)")
                continue

            # ------- Creation Folder -------

            qs_folder_list = self.creation.folder_line(value=code, description=desc, tooltips=False)

            # ------- Append -------

            if parent_id == 0:
                self.cat_model.appendRow(qs_folder_list)
                self.folder_dict[my_id] = qs_folder_list[col_cat_value]
                continue

            qs_current = self.folder_dict.get(parent_id)

            if not isinstance(qs_current, Folder):
                print("convert_manage -- ConvertMXDB -- get_folders -- parent_id not found")
                self.errors_list.append("get_folders -- parent_id not found")
                continue

            qs_current.appendRow(qs_folder_list)
            self.folder_dict[my_id] = qs_folder_list[col_cat_value]
            continue

    def get_children(self, cursor):

        query = """SELECT UM1, PREZZO1, DESCRIZIONE1, PUNTARTPADRE, PUNT_LIVELLOART, CODICEARTICOLO_NF 
                   FROM FElencoPrezzi ORDER BY CODICEARTICOLO_NF"""

        unit_col = 0
        price_col = 1
        desc_col = 2
        component_parent_col = 3
        material_parent_col = 4
        code_col = 5

        cursor.execute(query)

        rows = cursor.fetchall()

        qs_material = None
        material_desc = ""

        for row in rows:

            # ----------- Code -----------

            code = row[code_col]

            if not isinstance(code, str):
                print("convert_manage -- ConvertMXDB -- get_children -- not isinstance(code, str)")
                self.errors_list.append("get_children -- not isinstance(code, str)")
                continue

            code = code.strip()

            # ----------- Description -----------

            desc = row[desc_col]

            if not isinstance(desc, str):
                desc = ""
            else:
                desc = desc.strip()

            # ----------- Material -----------

            component_parent_id = row[component_parent_col]

            if not isinstance(component_parent_id, int) or component_parent_id == 0:

                material_parent_id = row[material_parent_col]

                if not isinstance(material_parent_id, int):
                    print("convert_manage -- ConvertMXDB -- get_children -- not isinstance(material_parent_id, int)")
                    self.errors_list.append("get_children -- not isinstance(material_parent_id, int)")
                    continue

                if material_parent_id < 1:
                    continue

                qs_parent = self.folder_dict.get(material_parent_id)

                if not isinstance(qs_parent, Folder):
                    print("convert_manage -- ConvertMXDB -- get_children -- not isinstance(qs_parent, Folder)")
                    self.errors_list.append("get_children -- not isinstance(qs_parent, Folder)")
                    continue

                qs_material_list = self.creation.material_line(value=code, description=desc, tooltips=False)

                qs_parent.appendRow(qs_material_list)

                qs_material = qs_material_list[col_cat_value]
                material_desc = desc

                continue

            # ----------- Component -----------

            if not isinstance(qs_material, Material):
                print("convert_manage -- ConvertMXDB -- get_children -- not isinstance(qs_material, Material)")
                self.errors_list.append("get_children -- not isinstance(qs_material, Material)")
                continue

            # ----------- Description -----------

            desc = f"{material_desc} {desc}"

            # ----------- Unit -----------

            unit = row[unit_col]

            if not isinstance(unit, str):
                unit = ""
            else:
                unit = unit.strip()

            # ----------- formula -----------

            if unit in self.unit_dict:
                formula = self.unit_dict[unit]

            else:
                formula = self.allplan.recherche_formule_defaut(unit=unit)

                self.unit_dict[unit] = formula

            # ----------- Price -----------

            price = row[price_col]

            if not isinstance(price, float):
                price = "0.00"
            else:
                price = f"{price:.2f}"

            qs_component_list = self.creation.component_line(value=code, description=desc, tooltips=False)

            qs_component: Component = qs_component_list[col_cat_value]

            qs_component.appendRow(self.creation.attribute_line(value=unit, number_str="202"))

            qs_component.appendRow(self.creation.attribute_line(value=price, number_str="203"))

            qs_component.appendRow(self.allplan.creation.attribute_line(value=formula, number_str="267"))

            qs_material.appendRow(qs_component_list)

    @staticmethod
    def a___________________end______():
        pass


def a___________________nevaris______():
    pass


class ConvertNevarisXml(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

        self.level_style_dict = {"NEVARIS_STYLE_H1": 0,
                                 "NEVARIS_STYLE_H2": 1,
                                 "NEVARIS_STYLE_H3": 2,
                                 "NEVARIS_STYLE_H4": 3,
                                 "NEVARIS_STYLE_H5": 4,
                                 "NEVARIS_STYLE_H6": 5,
                                 "NEVARIS_STYLE_Default": -1}

        self.code_index = 1
        self.type_index = 2
        self.desc_index = 3
        self.unit_index = 5

        self.root_code = "Raumelement"
        self.element_code = "Element"
        self.component_code = "Position"
        self.link_code = "Link"

        # --------------
        # Variables catalog datas
        # --------------

        self.material_list = list()
        self.material_upper_list = list()
        self.link_list = list()
        self.material_with_link_list = list()

        # --------------

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        self.start_loading()

        # --------------
        # Read / Parse file
        # --------------

        try:

            root = xml_load_root(file_path=self.file_path)

            if not isinstance(root, etree._Element):
                print(f"convert_manage -- ConvertNevarisXml -- run -- not isinstance(workbook, root, etree._Element)")
                self.errors_list.append("run -- not isinstance(workbook, root, etree._Element)")
                self.end_loading()
                return False

            # --------------
            # Search rows
            # --------------

            rows = root.iter("{urn:schemas-microsoft-com:office:spreadsheet}Row")

            # --------------
            # Variables
            # --------------

            rows_list = list()

            cell_datas_previous = dict()

            # --------------
            # Read all rows
            # --------------

            for row in rows:

                # --------------
                # Get style
                # --------------

                style_current = row.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}StyleID', 'Default')

                style_index = self.level_style_dict.get(style_current, -2)

                if style_index == -2:
                    print("convert_manage -- ConvertNevaris -- parse_excel_xml -- style_index not found: "
                          f"({style_current}")
                    self.errors_list.append("parse_excel_xml -- style_index not found")
                    continue

                # --------------
                # Get cell datas
                # --------------

                cell_datas = self.get_cells_datas(row=row, style_index=style_index,
                                                  cell_datas_previous=cell_datas_previous)

                if len(cell_datas) == 0:
                    continue

                rows_list.append(cell_datas)

                cell_datas_previous = cell_datas

            # --------------
            # Hierarchy_creation
            # --------------

            qs_levels = {0: self.cat_model.invisibleRootItem()}
            qs_folder = None
            qs_material = None

            qs_folder_dict = dict()

            qs_material_dict = dict()

            for cell_datas in rows_list:

                # --------------
                # Get Datas
                # --------------

                code_current: str = cell_datas["code"]
                type_current: str = cell_datas["type"]
                desc_current: str = cell_datas["desc"]
                unit_current: str = cell_datas["unit"]
                level_index: int = cell_datas["level"]

                # --------------
                # Get parent
                # --------------

                if type_current == folder_code:
                    qs_parent = qs_levels.get(level_index - 1, None)

                elif type_current == material_code:
                    qs_parent = qs_folder

                else:
                    qs_parent = qs_material

                # --------------
                # Check parent
                # --------------

                if not isinstance(qs_parent, QStandardItem):
                    print("convert_manage -- ConvertNevaris -- parse_excel_xml -- "
                          "not isinstance(qs_parent, QStandardItem)")
                    self.errors_list.append("parse_excel_xml -- not isinstance(qs_parent, QStandardItem)")
                    continue

                # --------------
                # Folder creation
                # --------------

                if type_current == folder_code:
                    qs_list = self.creation.folder_line(value=code_current,
                                                        description=desc_current,
                                                        tooltips=self.conversion)

                    qs_parent.appendRow(qs_list)

                    # ----

                    qs_folder = qs_list[0]
                    id_qs = id(qs_folder)

                    # ----

                    qs_levels[level_index] = qs_folder
                    qs_folder_dict[id_qs] = qs_folder

                    continue

                # --------------
                # Material creation
                # --------------

                if type_current == material_code:

                    used_by_links = self.link_list.count(code_current)

                    qs_list = self.creation.material_line(value=code_current,
                                                          description=desc_current,
                                                          used_by_links=used_by_links)

                    # ----

                    qs_material = qs_list[0]

                    # -----

                    id_parent_qs = id(qs_parent)

                    if id_parent_qs not in qs_material_dict:

                        qs_material_dict[id_parent_qs] = [qs_list]

                    else:

                        qs_material_list = qs_material_dict[id_parent_qs]
                        qs_material_list.append(qs_list)

                    continue

                # --------------
                # Component creation
                # --------------

                if type_current == component_code:

                    qs_list = self.creation.component_line(value=code_current, description=desc_current)

                    qs_parent.appendRow(qs_list)

                    # -----

                    if unit_current != "":
                        qs_value = qs_list[0]
                        qs_value.appendRow(self.creation.attribute_line(value=unit_current, number_str="202"))

                    continue

                # --------------
                # Link creation
                # --------------

                if type_current == link_code:

                    if code_current not in self.material_list:
                        self.errors_list.append(f"Link : {code_current} est orphenlin")
                        continue

                    # -----

                    qs_list = self.creation.link_line(value=code_current, description=desc_current)

                    qs_parent.appendRow(qs_list)

                    continue

                # --------------
                # Error creation
                # --------------

                print("convert_manage -- ConvertNevaris -- parse_excel_xml -- error type element")
                self.errors_list.append("parse_excel_xml -- error type element")

            # --------------
            # Material creation
            # --------------

            for id_folder, qs_material_list in qs_material_dict.items():

                if not isinstance(qs_material_list, list):
                    print("convert_manage -- ConvertNevaris -- parce_excel_xml -- "
                          "not isinstance(not isinstance(qs_material_list, list)")
                    continue

                qs_parent = qs_folder_dict.get(id_folder, None)

                if not isinstance(qs_parent, Folder):
                    print("convert_manage -- ConvertNevaris -- parce_excel_xml -- not isinstance(qs_parent, Folder)")
                    self.errors_list.append("parse_excel_xml -- not isinstance(qs_parent, Folder)")
                    continue

                parent_row_count = qs_parent.rowCount()

                if parent_row_count != 1:

                    material_row_first = qs_material_list[0]

                    if not isinstance(material_row_first, list):
                        print("convert_manage -- ConvertNevaris -- parce_excel_xml -- "
                              "not isinstance(material_row_first, list)")
                        self.errors_list.append("parse_excel_xml -- not isinstance(material_row_first, list)")
                        continue

                    if len(material_row_first) < 2:
                        print("convert_manage -- ConvertNevaris -- parce_excel_xml -- len(material_row_first) < 2")
                        self.errors_list.append("parse_excel_xml -- len(material_row_first) < 2")
                        continue

                    qs_code: Material = material_row_first[0]
                    qs_desc: Info = material_row_first[1]

                    material_text = qs_code.text()
                    material_desc = qs_desc.text()

                    # -------------

                    qs_list = self.creation.folder_line(value=material_text,
                                                        description=material_desc,
                                                        tooltips=self.conversion)

                    qs_parent = qs_list[0]

                    # -------------

                    qs_parent.appendRow(qs_list)

                for qs_material_row in qs_material_list:

                    if not isinstance(qs_material_row, list):
                        print("convert_manage -- ConvertNevaris -- parce_excel_xml -- "
                              "not isinstance(qs_material_row, list")
                        self.errors_list.append("parse_excel_xml -- not isinstance(qs_material_row, list)")
                        continue

                    qs_parent.appendRow(qs_material_row)

        except Exception as error:
            print(f"convert_manage -- ConvertNevaris -- parse_excel_xml -- error : {error}")
            self.errors_list.append(f"parse_excel_xml -- {error}")
            self.end_loading()
            return False

        self.end_loading()
        return True

    @staticmethod
    def a___________________tools______():
        pass

    def get_cells_datas(self, row, style_index: int, cell_datas_previous: dict) -> dict:

        cell_index = 1
        code_current = ""
        type_current = ""
        desc_current = ""

        cell_datas = {"code": "",
                      "type": type_current,
                      "desc": desc_current,
                      "unit": "",
                      "level": style_index}

        for cell in row:

            # --------------
            # Get if new index (cell empty)
            # --------------

            new_index = cell.attrib.get('{urn:schemas-microsoft-com:office:spreadsheet}Index', None)

            if isinstance(new_index, str):

                try:

                    new_index_int = int(new_index)

                    cell_index = new_index_int

                except Exception as error:
                    print(f"convert_manage -- ConvertNevaris -- get_cells_datas -- error : {error}")
                    self.errors_list.append(f"get_cells_datas -- {error}")
                    pass

            # --------------
            # Get code
            # --------------

            if cell_index == self.code_index:
                code_current = self.get_datas(cell=cell)
                cell_datas["code"] = code_current

                cell_index += 1
                continue

            # --------------
            # Get type
            # --------------

            if cell_index == self.type_index:

                type_current = self.get_datas(cell=cell)

                type_previous = cell_datas_previous.get("type", folder_code)

                # -------------
                # Folder / Element
                # -------------

                if type_current == self.element_code:
                    cell_index += 1
                    cell_datas["type"] = folder_code
                    continue

                if type_current != self.component_code and type_current != self.link_code:
                    return dict()

                # -------------
                code_previous = cell_datas_previous.get("code", "")
                # -------------

                # -------------
                # Component / Link
                # -------------

                if type_previous == component_code or type_previous == link_code:
                    pass

                # -------------
                # Material
                # -------------

                elif type_previous == folder_code:

                    type_previous = material_code

                    cell_datas_previous["type"] = type_previous

                    # -----------------

                    if code_previous in self.material_list:
                        code_previous_tps = find_new_title(base_title=code_previous,
                                                           titles_list=self.material_upper_list)

                        if code_previous != code_previous_tps:

                            self.errors_list.append(f"Material : {code_previous} a été renommé : {code_previous_tps}")

                            cell_datas_previous["code"] = code_previous_tps

                        else:

                            cell_datas_previous["code"] = code_previous

                    self.material_list.append(code_previous)
                    self.material_list.append(code_previous.upper())

                # -------------
                # Error
                # -------------

                else:
                    print(f"convert_manage -- ConvertNevaris -- get_cells_datas -- bad parent - {type_previous}")
                    self.errors_list.append(f"get_cells_datas -- bad parent - {type_previous}")
                    return dict()

                # -------------
                # Component
                # -------------

                if type_current == self.component_code:
                    cell_datas["type"] = component_code

                    cell_index += 1
                    continue

                # -------------
                # Link
                # -------------

                cell_datas["type"] = link_code
                self.link_list.append(code_current)

                if code_previous not in self.material_with_link_list:
                    self.material_with_link_list.append(code_previous.upper())

                cell_index += 1
                continue

            # --------------
            # Get description
            # --------------

            if cell_index == self.desc_index:

                desc_current = self.get_datas(cell=cell)
                cell_datas["desc"] = desc_current

                if type_current == self.element_code:

                    if code_current == "":
                        cell_datas["code"] = desc_current

                    return cell_datas

                if type_current == self.root_code:
                    return dict()

                cell_index += 1
                continue

            # --------------
            # Get unit
            # --------------

            if cell_index == self.unit_index:
                cell_datas["unit"] = self.get_datas(cell=cell)

                return cell_datas

            cell_index += 1

        return cell_datas

    def get_datas(self, cell):

        ele = cell.find('ss:Data', namespaces=self.ns)

        if ele is None:
            return

        value = ele.text

        if not isinstance(value, str):
            return ""

        return value

    @staticmethod
    def a___________________end______():
        pass


class ConvertNevarisExcel(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.code_index = 0
        self.type_index = 1
        self.desc_index = 2
        self.unit_index = 3
        self.text_index = 4
        self.number_index = 5
        self.formula_index = 8

        self.root_code = "Raumelement"
        self.element_code = "Element"
        self.component_code = "Position"
        self.link_code = "Link"

        self.exlude_list = ["Hinweis"]

        self.material_upper_list = list()

        # --------------

    @staticmethod
    def a___________________loading______():
        pass

    def run(self) -> bool:

        self.start_loading()

        # -----------------

        try:
            # --------------
            # Read / Parse file
            # --------------

            workbook = excel_load_workbook(file_path=self.file_path)

            if not isinstance(workbook, openpyxl.Workbook):
                print(f"convert_manage -- ConvertNevarisExcel -- run -- not isinstance(workbook, openpyxl.Workbook)")
                self.errors_list.append("run -- not isinstance(workbook, openpyxl.Workbook)")
                self.end_loading()
                return False

            sheet = workbook.active

            row_datas_previous = dict()

            row_list = list()

            for row in sheet.iter_rows(min_row=2, values_only=True):

                # ----------------------
                # read line
                # ----------------------

                row_datas = self.get_row_datas(row, row_datas_previous)

                if len(row_datas) == 0:
                    continue

                row_datas_previous = row_datas

                row_list.append(row_datas)

            qs_list = self.creation.folder_line(value=self.bdd_title, description="", tooltips=False)

            qs_root = qs_list[0]

            self.cat_model.appendRow(qs_list)

            qs_material = None

            for row_datas in row_list:

                type_current = row_datas.get("type", self.element_code)

                if type_current == self.element_code:
                    continue

                code_current = row_datas.get("code", "")

                if code_current == "":
                    continue

                desc_current = row_datas.get("desc", "")

                if type_current == material_code:
                    qs_list = self.creation.material_line(value=code_current, description=desc_current)

                    qs_root.appendRow(qs_list)

                    qs_material = qs_parent = qs_list[0]

                elif type_current == component_code:

                    qs_list = self.creation.component_line(value=code_current, description=desc_current)

                    qs_parent = qs_list[0]

                    qs_material.appendRow(qs_list)

                else:
                    continue

                if not isinstance(qs_parent, (Material, Component)):
                    continue

                attributes_dict = row_datas.get("attributes", dict())

                if not isinstance(attributes_dict, dict):
                    return

                for number, value in attributes_dict.items():
                    qs_parent.appendRow(self.creation.attribute_line(value=value, number_str=number))

        except Exception as error:
            print(f"convert_manage -- ConvertNevarisExcel -- run -- error : {error}")
            self.errors_list.append(f"parse_excel -- {error}")
            self.end_loading()
            return False

        self.end_loading()
        return True

    def get_row_datas(self, row: tuple, last_row: dict) -> dict:

        if len(row) < self.formula_index:
            print("convert_manage -- ConvertNevarisExcel -- get_row_datas -- len(row) < 4")
            self.errors_list.append(f"get_row_datas -- len(row) < self.formula_index")
            return dict()

        try:

            code_previous = last_row.get("code", "")
            type_previous = last_row.get("type", self.element_code)

            # ---------------------
            # Type
            # ---------------------

            type_current = row[self.type_index]

            if not isinstance(type_current, str):
                print("convert_manage -- ConvertNevarisExcel -- get_row_datas -- not isinstance(type_current, str)")
                self.errors_list.append(f"get_row_datas -- not isinstance(type_current, str)")
                return dict()

            if type_current == self.root_code:
                return dict()

            elif type_current == self.link_code or type_current == self.component_code:

                # --------------------------------
                # Change last Element to Material
                # --------------------------------

                if type_previous == self.element_code:

                    last_row["type"] = material_code

                    if code_previous.upper() not in self.material_upper_list:

                        self.material_upper_list.append(code_previous.upper())

                    else:

                        code_new = find_new_title(base_title=code_previous, titles_list=self.material_upper_list)

                        last_row["code"] = code_new

                        self.material_upper_list.append(code_new.upper())

                if type_current == self.link_code:
                    return dict()

                if type_current == self.component_code:
                    type_current = component_code

            elif type_current in self.exlude_list:
                return dict()

            elif type_current != self.element_code:
                print("convert_manage -- ConvertNevarisExcel -- get_row_datas -- type_current != self.element_code")
                self.errors_list.append(f"get_row_datas -- type_current != self.element_code")
                return dict()

            # ---------------------
            # Description
            # ---------------------

            desc_current = row[self.desc_index]

            if not isinstance(desc_current, str):
                desc_current = ""

            # ---------------------
            # code
            # ---------------------

            code_current = row[self.code_index]

            if not isinstance(code_current, str):
                code_current = desc_current

            row_datas = {"code": code_current,
                         "type": type_current,
                         "desc": desc_current}

            fusion = code_current == code_previous and type_current == type_previous

            if fusion:

                attributes_dict = last_row.get("attributes", dict())

            else:

                attributes_dict = dict()

            # ---------------------
            # Unit
            # ---------------------

            unit_current = row[self.unit_index]

            if not isinstance(unit_current, str):
                unit_current = ""

            if unit_current != "" and "202" not in attributes_dict:
                attributes_dict["202"] = unit_current

            # ---------------------
            # Text
            # ---------------------

            text_current: str = row[self.text_index]

            if not isinstance(text_current, str):
                text_current = ""

            if text_current != "" and "208" not in attributes_dict:
                attributes_dict["208"] = text_current

            # ---------------------
            # Number
            # ---------------------

            number_current: str = row[self.number_index]

            if not isinstance(number_current, str):
                number_current = ""

            # ---------------------
            # Formula
            # ---------------------

            formula_current = row[self.formula_index]

            if not isinstance(formula_current, str):
                formula_current = ""

            if number_current != "" and number_current not in attributes_dict:
                attributes_dict[number_current] = formula_current

            # ---------------------

            if not fusion:
                row_datas["attributes"] = attributes_dict

                return row_datas

            return dict()

        except Exception as error:
            print(f"convert_manage -- ConvertNevarisExcel -- get_row_datas --  error : {error}")
            self.errors_list.append(f"get_row_datas -- {error}")
            return dict()


def a___________________xpwe______():
    pass


class ConvertXpwe(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.qs_folder_0 = self.cat_model.invisibleRootItem()
        self.qs_folder_1 = None
        self.qs_folder_2 = None
        self.qs_material = None

        # --------------

    def run(self) -> bool:

        self.start_loading()

        try:

            root = xml_load_root(file_path=self.file_path)

            if not isinstance(root, etree._Element):
                print(f"convert_manage -- ConvertXpwe -- run -- not isinstance(root, etree._Element)")
                self.errors_list.append("run -- not isinstance(root, etree._Element)")
                self.end_loading()
                return False

            ep_item_list = root.findall(".//EPItem")

            for ep_item in ep_item_list:
                code_val = self.convert_data(ep_item.find('Tariffa'))

                if code_val == "":
                    print(f"convert_manage -- ConvertXpwe -- run -- code_val is empty")
                    self.errors_list.append("run -- code_val is empty")
                    continue

                desc_short_val = self.convert_data(ep_item.find("DesRidotta"))

                if len(code_val) < 5:
                    desc_long_val = unit_val = price_val = ""

                else:
                    desc_long_val = self.convert_data(ep_item.find("DesEstesa"))
                    unit_val = self.convert_data(ep_item.find("UnMisura"))

                    if len(code_val) > 7:
                        price_val = self.convert_data(ep_item.find("Prezzo1"))
                    else:
                        price_val = ""

                self.creation_add_item(code_val=code_val, desc_short_val=desc_short_val,
                                       desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

        except Exception as error:
            print(f"convert_manage -- ConvertXpwe -- run -- error : {error}")
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        # -----------------

        self.end_loading()
        return True

    def convert_data(self, value):
        try:
            value_str = value.text

            if isinstance(value_str, str):
                return value_str.strip()

        except Exception as error:
            print(f"convert_manage -- ConvertXpwe -- convert_data -- error : {error}")
            self.errors_list.append(f"convert_data -- {error}")
            pass

        return ""

    def creation_add_item(self, code_val: str, desc_short_val: str, desc_long_val="", unit_val="", price_val=""):

        code_count = len(code_val)

        if code_count == 1:

            if not isinstance(self.qs_folder_0, QStandardItem):
                print(f"convert_manage -- ConvertXpwe -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_0, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_0, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_1 = qs_list[col_cat_value]
            self.qs_folder_0.appendRow(qs_list)
            return True

        # -----------------

        if code_count < 5:

            if not isinstance(self.qs_folder_1, QStandardItem):
                print(f"convert_manage -- ConvertXpwe -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_1, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_1, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_2 = qs_list[col_cat_value]
            self.qs_folder_1.appendRow(qs_list)
            return True

        # -----------------

        elif code_count < 8:

            if not isinstance(self.qs_folder_2, QStandardItem):
                print(f"convert_manage -- ConvertXpwe -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_2, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_2, QStandardItem)")
                return False

            qs_list = self.allplan.creation.material_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_material = qs_list[col_cat_value]

            self.creation_add_attributes(qs_value=self.qs_material,
                                         desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

            self.qs_folder_2.appendRow(qs_list)
            return True

        else:

            if not isinstance(self.qs_material, QStandardItem):
                print(f"convert_manage -- ConvertXpwe -- creation_add_item -- "
                      f"not isinstance(self.qs_material, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_material, QStandardItem)")
                return False

            qs_list = self.allplan.creation.component_line(value=code_val, description=desc_short_val, tooltips=False)

            self.creation_add_attributes(qs_value=qs_list[0],
                                         desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

            self.qs_material.appendRow(qs_list)
            return True

    def creation_add_attributes(self, qs_value: QStandardItem, desc_long_val: str, unit_val: str, price_val: str):

        if not isinstance(qs_value, QStandardItem):
            print(f"convert_manage -- ConvertXpwe -- creation_add_attributes -- "
                  f"not isinstance(self.qs_value, QStandardItem)")
            self.errors_list.append("creation_add_attributes -- not isinstance(self.qs_value, QStandardItem)")
            return False

        if unit_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=unit_val, number_str="202",
                                                                    model_enumeration=self.allplan.model_units))

        if price_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=price_val, number_str="203"))

        if desc_long_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=desc_long_val, number_str="208"))

        if isinstance(qs_value, Component):
            formula = self.allplan.recherche_formule_defaut(unit=unit_val)
            qs_value.appendRow(self.allplan.creation.attribute_line(value=formula, number_str="267"))

        return True


def a___________________team_system______():
    pass


class ConvertTeamSystemXml(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.unit_dict = dict()
        self.ns = {"six": "six.xsd"}

        self.folder_len = 7
        self.material_len = 12

        self.qs_folder_0 = self.cat_model.invisibleRootItem()
        self.qs_folder_1 = None
        self.qs_folder_2 = None
        self.qs_folder_3 = None
        self.qs_material = None

        # --------------

    def run(self) -> bool:

        self.start_loading()

        try:
            root = xml_load_root(file_path=self.file_path)

            if not isinstance(root, etree._Element):
                print(f"convert_manage -- ConvertTeamSystemXml -- run -- not isinstance(root, etree._Element)")
                self.errors_list.append("run -- not isinstance(root, etree._Element)")
                self.end_loading()
                return False

            # -----------------

            unit_ele_list = root.findall(".//six:unitaDiMisura", namespaces=self.ns)

            self.convert_get_unit_dict(unit_ele_list=unit_ele_list)

            # -----------------

            prodotto_list = root.findall(".//six:prodotto", namespaces=self.ns)

            # -----------------

            for prodotto in prodotto_list:

                code_val, unit_val = self.convert_get_code(prodotto=prodotto)

                if code_val == "":
                    print(f"convert_manage -- ConvertTeamSystemXml -- run -- code_val is empty")
                    self.errors_list.append("run -- code_val is empty")
                    continue

                code_len = len(code_val)

                if code_len == 0:
                    print(f"convert_manage -- ConvertTeamSystemXml -- run -- code_len == 0")
                    self.errors_list.append("run -- code_len == 0")
                    continue

                # -----------------

                desc_short_val, desc_long_val = self.convert_get_desc(value=prodotto,
                                                                      get_desc_long=code_len >= self.material_len)

                # -----------------

                if code_len <= self.folder_len:
                    self.creation_add_item(code_val=code_val, desc_short_val=desc_short_val)
                    continue

                # -----------------

                if code_len <= self.material_len:
                    self.creation_add_item(code_val=code_val, desc_short_val=desc_short_val,
                                           desc_long_val=desc_long_val)
                    continue

                # -----------------

                price_val = self.convert_get_price(value=prodotto)

                self.creation_add_item(code_val=code_val, desc_short_val=desc_short_val,
                                       desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

        except Exception as error:
            print(f"convert_manage -- ConvertXpwe -- run -- error : {error}")
            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        # -----------------

        self.end_loading()
        return True

    def creation_add_item(self, code_val: str, desc_short_val: str, desc_long_val="",
                          unit_val="", price_val="") -> bool:

        code_count = len(code_val)

        if code_count == 1:

            if not isinstance(self.qs_folder_0, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystem -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_0, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_1, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_1 = qs_list[col_cat_value]
            self.qs_folder_0.appendRow(qs_list)
            return True

        # -----------------

        if code_count < 5:

            if not isinstance(self.qs_folder_1, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystem -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_1, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_1, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_2 = qs_list[col_cat_value]
            self.qs_folder_1.appendRow(qs_list)
            return True

        # -----------------

        elif code_count <= self.folder_len:

            if not isinstance(self.qs_folder_2, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystem -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_2, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_2, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_3 = qs_list[col_cat_value]
            self.qs_folder_2.appendRow(qs_list)
            return True

        # -----------------

        elif code_count <= self.material_len:

            if not isinstance(self.qs_folder_3, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystem -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_3, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_3, QStandardItem)")
                return False

            qs_list = self.allplan.creation.material_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_material = qs_list[col_cat_value]

            self.creation_add_attributes(qs_value=self.qs_material,
                                         desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

            self.qs_folder_3.appendRow(qs_list)
            return True

        else:

            if not isinstance(self.qs_material, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystem -- creation_add_item -- "
                      f"not isinstance(self.qs_material, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_material, QStandardItem)")
                return False

            qs_list = self.allplan.creation.component_line(value=code_val, description=desc_short_val)

            self.creation_add_attributes(qs_value=qs_list[0],
                                         desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

            self.qs_material.appendRow(qs_list)
            return True

    def creation_add_attributes(self, qs_value: QStandardItem, desc_long_val: str,
                                unit_val: str, price_val: str) -> bool:

        if not isinstance(qs_value, QStandardItem):
            print(f"convert_manage -- ConvertTeamSystem -- creation_add_attributes -- "
                  f"not isinstance(self.qs_value, QStandardItem)")
            self.errors_list.append("creation_add_attributes -- not isinstance(self.qs_value, QStandardItem)")
            return False

        if unit_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=unit_val, number_str="202",
                                                                    model_enumeration=self.allplan.model_units))

        if price_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=price_val, number_str="203"))

        if desc_long_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=desc_long_val, number_str="208"))

        if isinstance(qs_value, Component):
            formula = self.allplan.recherche_formule_defaut(unit=unit_val)
            qs_value.appendRow(self.allplan.creation.attribute_line(value=formula, number_str="267"))

        return True

    def convert_get_unit_dict(self, unit_ele_list: list) -> bool:

        if not isinstance(unit_ele_list, list):
            print("convert_manage -- ConvertTeamSystem -- convert_get_unit_dict -- not isinstance(unit_ele, list)")
            self.errors_list.append("convert_get_unit_dict -- not isinstance(unit_ele_list, list)")
            return False

        for unit_ele in unit_ele_list:

            if unit_ele is None:
                print("convert_manage -- ConvertTeamSystem -- convert_get_unit_dict -- unit_ele is None")
                self.errors_list.append("convert_get_unit_dict -- not isinstance(unit_ele is None)")
                continue

            unit_id = unit_ele.get("unitaDiMisuraId")

            if not isinstance(unit_id, str):
                print("convert_manage -- ConvertTeamSystem -- convert_get_unit_dict -- not isinstance(unit_id, str)")
                self.errors_list.append("convert_get_unit_dict -- not isinstance(unit_id, str)")
                continue

            unit_val = unit_ele.get("simbolo")

            if not isinstance(unit_id, str):
                print("convert_manage -- ConvertTeamSystem -- convert_get_unit_dict -- not isinstance(unit_id, str)")
                self.errors_list.append("convert_get_unit_dict -- not isinstance(unit_id, str)")
                continue

            self.unit_dict[unit_id] = unit_val

        return True

    def convert_get_code(self, prodotto) -> tuple:

        try:

            code_val = prodotto.get("prdId")

            if not isinstance(code_val, str):
                code_val = ""

            # ----------------

            if len(code_val) < self.material_len:
                return code_val, ""

            unit_val = prodotto.get("unitaDiMisuraId")

            if not isinstance(unit_val, str):
                unit_val = ""
            else:
                unit_val = self.unit_dict.get(unit_val, "")

            return code_val, unit_val

        except Exception as error:
            print(f"convert_manage -- ConvertTeamSystem -- convert_get_code -- error : {error}")
            self.errors_list.append(f"convert_get_code -- {error}")
            pass

        return "", ""

    def convert_get_desc(self, value, get_desc_long=False) -> tuple:

        try:
            sub_value = value.find("six:prdDescrizione", namespaces=self.ns)

            if sub_value is None:
                return "", ""

            # ----------------

            desc_short_val = sub_value.get("breve")

            if not isinstance(desc_short_val, str):
                desc_short_val = ""

            # ----------------

            if not get_desc_long:
                return desc_short_val, ""

            desc_long_val = sub_value.get("estesa")

            if not isinstance(desc_long_val, str):
                desc_long_val = ""

            return desc_short_val, desc_long_val

        except Exception as error:
            print(f"convert_manage -- ConvertTeamSystem -- convert_get_desc -- error : {error}")
            self.errors_list.append(f"convert_get_desc -- {error}")
            pass

        return "", ""

    def convert_get_price(self, value) -> str:

        try:
            sub_value = value.find("six:prdQuotazione", namespaces=self.ns)

            if sub_value is None:
                return ""

            # ----------------

            price_val = sub_value.get("valore")

            if not isinstance(price_val, str):
                return ""

            return price_val

        except Exception as error:
            print(f"convert_manage -- ConvertTeamSystem -- convertconvert_get_price_get_desc -- error : {error}")
            self.errors_list.append(f"convert_get_price -- {error}")
            pass

        return ""


class ConvertTeamSystemXlsx(ConvertTemplate):

    def __init__(self, allplan: AllplanDatas, file_path: str, bdd_title: str, conversion=False):
        super().__init__(allplan, file_path, bdd_title, conversion)

        # --------------
        # Variables
        # --------------

        self.folder_len = 8
        self.material_len = 13

        # --------------

        self.code_col = 0
        self.desc_short_col = 1
        self.desc_long_col = 2
        self.unit_col = 4
        self.price_col = 5

        self.last_col = 6

        self.qs_folder_0 = self.cat_model.invisibleRootItem()
        self.qs_folder_1 = None
        self.qs_folder_2 = None
        self.qs_folder_3 = None
        self.qs_material = None

        # --------------

    def run(self) -> bool:

        self.start_loading()

        workbook = excel_load_workbook(file_path=self.file_path)

        if not isinstance(workbook, openpyxl.Workbook):
            print(f"convert_manage -- ConvertTeamSystemXlsx -- run -- not isinstance(workbook, openpyxl.Workbook)")
            self.errors_list.append("not isinstance(workbook, openpyxl.Workbook)")
            self.end_loading()
            return False

        try:
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=3, values_only=True):

                datas = self.creation_get_datas(row=row)

                if datas is None:
                    # print("convertmanage -- ConvertTeamSystemXlsx -- run -- datas is None")
                    continue

                code_val, desc_short_val, desc_long_val, unit_val, price_val = datas

                code_len = len(code_val)

                if code_len == 0:
                    print(f"convert_manage -- ConvertTeamSystemXlsx -- run -- code_len == 0")
                    self.errors_list.append("run -- code_len == 0")
                    continue

                # -----------------

                if code_len <= self.folder_len:
                    self.creation_add_item(code_val=code_val, desc_short_val=desc_short_val)
                    continue

                # -----------------

                if code_len <= self.material_len:
                    self.creation_add_item(code_val=code_val, desc_short_val=desc_short_val,
                                           desc_long_val=desc_long_val)
                    continue

                # -----------------

                self.creation_add_item(code_val=code_val, desc_short_val=desc_short_val,
                                       desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

        except Exception as error:

            print(f"convert_manage -- ConvertXpwe -- run -- error : {error}")

            # -----------------

            self.errors_list.append(f"run -- {error}")
            self.end_loading()
            return False

        # -----------------

        self.end_loading()
        return True

    def creation_add_item(self, code_val: str, desc_short_val: str, desc_long_val="",
                          unit_val="", price_val="") -> bool:

        code_count = len(code_val)

        if code_count == 1:

            if not isinstance(self.qs_folder_0, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystemXlsx -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_0, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_0, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_1 = qs_list[col_cat_value]
            self.qs_folder_0.appendRow(qs_list)
            return True

        # -----------------

        if code_count < 5:

            if not isinstance(self.qs_folder_1, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystemXlsx -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_1, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_1, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_2 = qs_list[col_cat_value]
            self.qs_folder_1.appendRow(qs_list)
            return True

        # -----------------

        elif code_count <= self.folder_len:

            if not isinstance(self.qs_folder_2, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystemXlsx -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_2, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_2, QStandardItem)")
                return False

            qs_list = self.allplan.creation.folder_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_folder_3 = qs_list[col_cat_value]
            self.qs_folder_2.appendRow(qs_list)
            return True

        # -----------------

        elif code_count <= self.material_len:

            if not isinstance(self.qs_folder_3, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystemXlsx -- creation_add_item -- "
                      f"not isinstance(self.qs_folder_3, QStandardItem)")
                self.errors_list.append("creation_add_item -- not isinstance(self.qs_folder_3, QStandardItem)")
                return False

            qs_list = self.allplan.creation.material_line(value=code_val, description=desc_short_val, tooltips=False)

            self.qs_material = qs_list[col_cat_value]

            self.creation_add_attributes(qs_value=self.qs_material,
                                         desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

            self.qs_folder_3.appendRow(qs_list)
            return True

        else:

            if not isinstance(self.qs_material, QStandardItem):
                print(f"convert_manage -- ConvertTeamSystemXlsx -- creation_add_item -- "
                      f"not isinstance(self.qs_material, QStandardItem)")

                self.errors_list.append("creation_add_item -- not isinstance(self.qs_material, QStandardItem)")
                return False

            qs_list = self.allplan.creation.component_line(value=code_val, description=desc_short_val)

            self.creation_add_attributes(qs_value=qs_list[0],
                                         desc_long_val=desc_long_val, unit_val=unit_val, price_val=price_val)

            self.qs_material.appendRow(qs_list)
            return True

    def creation_add_attributes(self, qs_value: QStandardItem, desc_long_val: str,
                                unit_val: str, price_val: str) -> bool:

        if not isinstance(qs_value, QStandardItem):
            print(f"convert_manage -- ConvertTeamSystemXlsx -- creation_add_attributes -- "
                  f"not isinstance(qs_value, QStandardItem)")
            self.errors_list.append("creation_add_attributes -- not isinstance(qs_value, QStandardItem)")
            return False

        if unit_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=unit_val, number_str="202",
                                                                    model_enumeration=self.allplan.model_units))

        if price_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=price_val, number_str="203"))

        if desc_long_val != "":
            qs_value.appendRow(self.allplan.creation.attribute_line(value=desc_long_val, number_str="208"))

        if isinstance(qs_value, Component):
            formula = self.allplan.recherche_formule_defaut(unit=unit_val)
            qs_value.appendRow(self.allplan.creation.attribute_line(value=formula, number_str="267"))

        return True

    def creation_get_datas(self, row) -> tuple | None:

        if len(row) < self.last_col:
            print(f"convert_manage -- ConvertTeamSystemXlsx -- creation_get_datas -- len(row) < self.last_col")
            self.errors_list.append("creation_get_datas -- len(row) < self.last_col")
            return

        try:

            code_val: str = row[self.code_col]

            if not isinstance(code_val, str):
                return

            code_len = len(code_val)

            # ----------

            desc_short_val: str = row[self.desc_short_col]

            if not isinstance(desc_short_val, str):
                desc_short_val = ""

            # ----------

            if code_len <= self.folder_len:
                return code_val, desc_short_val, "", "", ""

            # ----------

            desc_long_val = row[self.desc_long_col]

            if not isinstance(desc_long_val, str):
                desc_long_val = ""

            # ----------

            if code_len <= self.material_len:
                return code_val, desc_short_val, desc_long_val, "", ""

            # ----------

            unit_val = row[self.unit_col]

            if not isinstance(unit_val, str):
                unit_val = ""

            # ----------

            price_val = row[self.price_col]

            if isinstance(price_val, float) or isinstance(price_val, int):
                price_val = f"{price_val:.2f}"

            elif not isinstance(price_val, str):
                price_val = ""

            return code_val, desc_short_val, desc_long_val, unit_val, price_val

        # ----------

        except Exception as error:
            print(f"convertmanage -- ConvertTeamSystemXlsx -- run -- error : {error}")
            self.errors_list.append(f"creation_get_datas -- {error}")


def a____export___export___export___export___export___():
    pass


class ExportExcel(QObject):

    def __init__(self, allplan: AllplanDatas, model_cat: QStandardItemModel, chemin_fichier: str):
        super().__init__()

        self.allplan = allplan

        self.model_cat = model_cat

        self.chemin_fichier = chemin_fichier
        self.chemin_dossier = find_folder_path(chemin_fichier)

        self.datas_attributs = dict()
        self.dict_index_attributs = dict()
        self.datas = list()

        self.col_id = 3

        self.export_model()

    @staticmethod
    def a___________________attributs______():
        """ Partie réservée à la recherche des données"""
        pass

    def gestion_attributs(self, qs_parent: MyQstandardItem, index_enfant: int):

        qs_value: MyQstandardItem = qs_parent.child(index_enfant, col_cat_value)

        if not isinstance(qs_value, Attribute):
            return None

        qs_number: MyQstandardItem = qs_parent.child(index_enfant, col_cat_number)

        if not isinstance(qs_number, Info):
            return None

        number_str = qs_number.text()

        if number_str == attribut_default_obj.current:
            return ["", ""]

        valeur_attribut: str = qs_value.text()

        if not isinstance(valeur_attribut, str):
            print("convert_manage -- ExportExcel -- gestion_attributs -- not isinstance(valeur_attribut, str)")
            return None

        attribute_obj = self.allplan.attributes_dict.get(number_str)

        if not isinstance(attribute_obj, AttributeDatas):
            print("convert_manage -- ExportExcel -- gestion_attributs -- not isinstance(valeur_attribut, str)")
            return None

        type_ele2: str = attribute_obj.option

        if type_ele2 == code_attr_combo_int:

            qs_index: MyQstandardItem = qs_parent.child(index_enfant, col_cat_index)

            if not isinstance(qs_index, Info):
                print("convert_manage -- ExportExcel -- gestion_attributs -- not isinstance(qs_index, Info)")
                return None

            valeur_attribut = qs_index.text()

            if not isinstance(valeur_attribut, str):
                print("convert_manage -- ExportExcel -- gestion_attributs -- not isinstance(valeur_attribut, str)")
                return None

        elif type_ele2 == code_attr_chk:
            if valeur_attribut == "true":
                valeur_attribut = "1"
            else:
                valeur_attribut = "0"

        elif type_ele2 in [code_attr_formule_str, code_attr_formule_int, code_attr_formule_float]:

            if "\n" in valeur_attribut:
                valeur_attribut = valeur_attribut.replace("\n", "")

            try:
                numero_int = int(number_str)

                if 1999 < numero_int < 12000:
                    valeur_attribut = "1"

            except ValueError:
                pass

        return [number_str, valeur_attribut]

    @staticmethod
    def a___________________model______():
        """ Partie réservée à la recherche des données"""
        pass

    def export_model(self):

        self.datas = [["Type", "Parent", "Attribute", "Value"]]

        self.export_model_creation(self.model_cat.invisibleRootItem())
        self.export_model_to_excel()

    def export_model_creation(self, qs_parent: MyQstandardItem):

        if not isinstance(qs_parent, QStandardItem):
            print("convert_manage -- ExportExcel -- export_model_creation -- not isinstance(qs_parent, MyQs)")
            return None

        nb_enfant = qs_parent.rowCount()

        if qs_parent == self.model_cat.invisibleRootItem():
            texte_parent = ""
        else:
            texte_parent = qs_parent.text()

        if nb_enfant == 0:
            return

        for index_enfant in range(nb_enfant):
            qs_val: MyQstandardItem = qs_parent.child(index_enfant, col_cat_value)

            if isinstance(qs_val, Attribute):
                self.export_model_attributs(qs_parent, index_enfant)
                continue

            texte_val: str = qs_val.text()

            if isinstance(qs_val, Link):
                self.datas.append(["Link", texte_parent, "", texte_val])
                continue

            if isinstance(qs_val, Folder):
                self.datas.append(["Folder", texte_parent, "", texte_val])

            elif isinstance(qs_val, Material):
                self.datas.append(["Material", texte_parent, "", texte_val])

            elif isinstance(qs_val, Component):
                self.datas.append(["Component", texte_parent, "", texte_val])

            if qs_val.hasChildren():
                self.export_model_creation(qs_val)

    def export_model_attributs(self, qs_parent: MyQstandardItem, index_enfant: int):

        attribut = self.gestion_attributs(qs_parent, index_enfant)

        if not isinstance(attribut, list):
            return

        numero, valeur_attribut = attribut

        texte_parent: str = qs_parent.text()
        self.datas.append(["Attribute", texte_parent, numero, valeur_attribut])

    def export_model_to_excel(self):

        wb = Workbook()
        sheet = wb.active

        try:
            for row in self.datas:
                sheet.append(row)

            wb.save(self.chemin_fichier)

        except Exception as erreur:
            print(f"{erreur}")

            msg(titre=application_title,
                message=self.tr("Une erreur est survenue."),
                icone_avertissement=True,
                details=f"{erreur}")

            return

        msg(titre=application_title,
            message=self.tr("L'export s'est correctement déroulé!"),
            icone_valide=True)

        open_file(self.chemin_fichier)


class ExportCatalog(QObject):

    def __init__(self, catalogue, allplan, file_path: str, brand: str):

        super().__init__()

        self.allplan: AllplanDatas = allplan

        self.catalogue = catalogue

        self.model = self.catalogue.cat_model

        self.file_path = file_path

        self.brand = brand

        self.sauvegarde_catalogue()

    def sauvegarde_catalogue(self):

        tps = time.perf_counter()

        root = etree.Element(self.brand)

        self.sauvegarde_hierarchie(self.model.invisibleRootItem(), root)

        a = self.tr("Une erreur est survenue.")

        try:
            catalogue = etree.tostring(root,
                                       pretty_print=True,
                                       xml_declaration=True,
                                       encoding='UTF-8').decode()

            print(f"ExportCatalog : {time.perf_counter() - tps}s")

        except Exception as erreur:
            msg(titre=application_title,
                message=f'{a} : {self.file_path}',
                icone_critique=True,
                details=f"{erreur}")
            return False

        try:

            with open(self.file_path, "w", encoding="utf_8_sig") as file:
                file.write(catalogue)

        except Exception as erreur:
            msg(titre=application_title,
                message=f'{a} : {self.file_path}',
                icone_critique=True,
                details=f"{erreur}")
            return False

    def sauvegarde_hierarchie(self, qs_parent: MyQstandardItem, racine: etree._Element):

        for index_row in range(qs_parent.rowCount()):

            qs_value = qs_parent.child(index_row, col_cat_value)

            if isinstance(qs_value, Attribute) or isinstance(qs_value, Link):
                continue

            qs_desc = qs_parent.child(index_row, col_cat_desc)

            if not isinstance(qs_desc, Info):
                print("convert_manage -- ExportCatalog -- sauvegarde_hierarchie -- not isinstance(qs_desc, Info)")
                continue

            description = qs_desc.text()

            if not isinstance(description, str):
                print("convert_manage -- ExportCatalog -- sauvegarde_hierarchie -- not isinstance(description, str)")
                continue

            if isinstance(qs_value, Folder):
                node = self.creation_dossier(qs=qs_value, racine=racine, description=description)

                if not qs_value.hasChildren():
                    continue

                self.sauvegarde_hierarchie(qs_parent=qs_value, racine=node)
                continue

            unit = self.get_unit(qs_value)

            if isinstance(qs_value, Material):
                # group = self.creation_ouvrage(qs=qs, description=description, unit=unit, node=racine)

                if not qs_value.hasChildren():
                    continue

                self.sauvegarde_hierarchie(qs_parent=qs_value, racine=racine)
                continue

            if isinstance(qs_value, Component):
                self.creation_composant(qs=qs_value, description=description, unit=unit, group=racine)
                continue

    @staticmethod
    def creation_dossier(qs: QStandardItem, description: str, racine: etree._Element):

        titre = qs.text()
        node = etree.SubElement(racine, "Folder", name=titre, description=description)

        return node

    def creation_ouvrage(self, qs: MyQstandardItem, description: str, unit: str, node: etree._Element):

        group = etree.SubElement(node, "Material", name=qs.text(), description=description, unit=unit)

        self.creation_attributs(qs=qs, definition=group)

        return group

    def creation_composant(self, qs: MyQstandardItem, description: str, unit: str, group: etree._Element):

        position = etree.SubElement(group, "Component", name=qs.text(), description=description, unit=unit)

        self.creation_attributs(qs=qs, definition=position)

    def creation_attributs(self, qs: MyQstandardItem, definition: etree._Element):

        nb_enfants = qs.rowCount()
        plume = True
        trait = True
        couleur = True

        for index_row in range(nb_enfants):

            qs_value_child = qs.child(index_row, col_cat_value)

            if not isinstance(qs_value_child, Attribute):
                print("convert_manage -- ExportCatalog -- creation_attributs -- "
                      "not isinstance(qs_child_value, Attribute)")
                return

            qs_number_child = qs.child(index_row, col_cat_number)

            if not isinstance(qs_number_child, Info):
                print("convert_manage -- ExportCatalog -- creation_attributs -- not isinstance(qs_child_number, Info)")
                return

            valeur = qs_value_child.text()
            number_str = qs_number_child.text()

            if number_str == "207" or number_str == "202":
                continue

            if number_str in liste_attributs_with_no_val_no_save and valeur == "":
                continue

            if number_str == "349":
                plume, trait, couleur = self.gestion_layer(valeur)

            if ((number_str == "346" and not plume) or (number_str == "345" and not trait) or
                    (number_str == "347" and not couleur)):
                continue

            attribute_obj = self.allplan.attributes_dict.get(number_str)

            if not isinstance(attribute_obj, AttributeDatas):
                print("convert_manage -- ExportCatalog -- creation_attributs -- "
                      "not isinstance(attribute_obj, AttributeDatas)")
                continue

            type_ele2: str = attribute_obj.option

            if type_ele2 == code_attr_combo_int:
                qs_child_index = qs.child(index_row, col_cat_index)

                if not isinstance(qs_child_index, Info):
                    print("convert_manage -- ExportCatalog -- creation_attributs -- "
                          "not isinstance(qs_child_index, Info))")
                    continue

                valeur = qs_child_index.text()

                if not isinstance(valeur, str):
                    print("convert_manage -- ExportCatalog -- creation_attributs -- not isinstance(valeur, str)")
                    continue

                etree.SubElement(definition, 'Attribute', id=number_str, value=valeur)
                continue

            if type_ele2 in [code_attr_formule_str, code_attr_formule_int, code_attr_formule_float]:

                if "\n" in valeur:
                    valeur = valeur.replace("\n", "")

                try:
                    numero_int = int(number_str)

                    if 1999 < numero_int < 12000:
                        valeur = "1"

                    else:

                        valeur = self.allplan.convertir_formule(valeur)

                except ValueError:
                    pass

                etree.SubElement(definition, 'Attribute', id=number_str, value=valeur)
                continue

            etree.SubElement(definition, 'Attribute', id=number_str, value=valeur)
        return

    @staticmethod
    def gestion_layer(numero_style: str):

        plume = True
        trait = True
        couleur = True

        if numero_style == "1":
            plume = False
            return plume, trait, couleur

        if numero_style == "2":
            trait = False
            return plume, trait, couleur

        if numero_style == "3":
            plume = False
            trait = False
            return plume, trait, couleur

        if numero_style == "4":
            couleur = False
            return plume, trait, couleur

        if numero_style == "5":
            plume = False
            couleur = False
            return plume, trait, couleur

        if numero_style == "6":
            trait = False
            couleur = False
            return plume, trait, couleur

        if numero_style == "7":
            plume = False
            trait = False
            couleur = False
            return plume, trait, couleur

        return plume, trait, couleur

    @staticmethod
    def get_unit(qs: MyQstandardItem):

        search = qs.get_attribute_value_by_number("202")

        if search is None:
            return ""

        return search


def a___________________end______():
    pass
